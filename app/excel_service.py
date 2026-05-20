from __future__ import annotations

import re
import shutil
import subprocess
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .database import EXPORT_DIR, TEMPLATE_DIR, get_conn

CHECK_MARK = "✓"
HEAD_DELEGATION_TEXT = "Đã ủy quyền. Đồng ý với nội dung kiểm tra."
CONCLUSION_TEXT = (
    "Ban Kiểm soát nội bộ thực hiện kiểm soát trước đối với hồ sơ nêu trên theo chức năng "
    "kiểm soát, thẩm tra và kiến nghị; không thay thế trách nhiệm lập hồ sơ của Phòng Tài chính "
    "kế toán hoặc đơn vị liên quan, không thay thế thẩm quyền quyết định, phê duyệt của Hội đồng "
    "thành viên hoặc Chủ tịch Hội đồng thành viên."
)

REVIEW_ITEM_MASTER = [
    "Căn cứ pháp lý của khoản thu, khoản chi",
    "Căn cứ nội bộ, thẩm quyền trình và phê duyệt",
    "Tính đầy đủ của hồ sơ, tài liệu kèm theo",
    "Tính hợp lệ, hợp pháp, thống nhất của chứng từ",
    "Sự phù hợp với nghị quyết, quyết định, kế hoạch, dự toán, hợp đồng, biên bản nghiệm thu hoặc hồ sơ liên quan",
    "Sự phù hợp về số tiền, nội dung, đối tượng, thời điểm thu, chi",
    "Việc tuân thủ phân công, phân nhiệm, bất kiêm nhiệm, ủy quyền",
    "Dấu hiệu xung đột lợi ích, ưu ái, thiếu minh bạch nếu có",
    "Dấu hiệu bất thường, vượt thẩm quyền, vượt định mức, vượt dự toán hoặc rủi ro khác",
    "Nội dung khác cần lưu ý",
]


def _norm(v: Any) -> str:
    return str(v or "").strip()


def _money(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(".", "").replace(",", "").replace(" ", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def supplier_match_key(value: Any) -> str:
    text = _norm(value).upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("Đ", "D")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _header_key(value: Any) -> str:
    return supplier_match_key(value)


def _find_header_columns(ws, expected_headers: set[str]) -> tuple[int, dict[str, int]]:
    max_scan_row = min(ws.max_row, 30)
    for row_idx in range(1, max_scan_row + 1):
        columns: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            key = _header_key(ws.cell(row_idx, col_idx).value)
            if key:
                columns[key] = col_idx

        if any(header in columns for header in expected_headers):
            return row_idx, columns

    return 0, {}


def _find_column(columns: dict[str, int], candidates: list[str]) -> int | None:
    normalized_candidates = [_header_key(candidate) for candidate in candidates]

    for candidate in normalized_candidates:
        if candidate in columns:
            return columns[candidate]

    for header, col_idx in columns.items():
        for candidate in normalized_candidates:
            if candidate and candidate in header:
                return col_idx

    return None


def parse_dntt_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []

    for r in range(1, ws.max_row + 1):
        supplier = _norm(ws.cell(r, 1).value)
        content = _norm(ws.cell(r, 2).value)
        amount = _money(ws.cell(r, 3).value)

        supplier_lower = supplier.lower()
        content_lower = content.lower()

        if supplier_lower in {"nhà cung cấp", "nha cung cap"}:
            continue
        if content_lower in {"nội dung", "noi dung"}:
            continue
        if not supplier and not content:
            continue
        if amount <= 0:
            continue

        rows.append(
            {
                "source_type": "DNTT",
                "supplier": supplier,
                "content": content,
                "document_amount": amount,
            }
        )

    return rows


def parse_von_tu_co_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []

    header_row, columns = _find_header_columns(
        ws,
        {
            "MA",
            "NHA CUNG CAP",
            "NOI DUNG",
            "SO TIEN",
        },
    )

    if header_row:
        code_col = _find_column(columns, ["Mã", "Ma", "Mã đối tượng", "Ma doi tuong"])
        supplier_col = _find_column(columns, ["Nhà cung cấp", "Nha cung cap", "Tên nhà cung cấp", "Ten nha cung cap"]) or 1
        content_col = _find_column(columns, ["Nội dung", "Noi dung", "Diễn giải", "Dien giai"]) or 2
        amount_col = _find_column(columns, ["Số tiền", "So tien", "Số tiền thanh toán", "So tien thanh toan"]) or 3
        start_row = header_row + 1
    else:
        code_col = None
        supplier_col = 1
        content_col = 2
        amount_col = 3
        start_row = 1

    for r in range(start_row, ws.max_row + 1):
        supplier_code = _norm(ws.cell(r, code_col).value) if code_col else ""
        supplier = _norm(ws.cell(r, supplier_col).value)
        content = _norm(ws.cell(r, content_col).value)
        amount = _money(ws.cell(r, amount_col).value)

        supplier_lower = supplier.lower()
        content_lower = content.lower()

        if supplier_lower in {"nhà cung cấp", "nha cung cap"}:
            continue
        if content_lower in {"nội dung", "noi dung"}:
            continue
        if not supplier and not content:
            continue
        if amount <= 0:
            continue

        rows.append(
            {
                "source_type": "VON_TU_CO",
                "supplier_code": supplier_code,
                "supplier": supplier,
                "content": content,
                "document_amount": amount,
            }
        )

    return rows


def parse_thcn_payable_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []

    for r in range(1, ws.max_row + 1):
        account_code = _norm(ws.cell(r, 6).value)
        supplier_code = _norm(ws.cell(r, 7).value)
        supplier_name = _norm(ws.cell(r, 10).value)
        ending_credit = _money(ws.cell(r, 24).value)

        if account_code != "331":
            continue
        if not supplier_name:
            continue

        rows.append(
            {
                "source_type": "THCN_PAYABLE",
                "supplier_code": supplier_code,
                "supplier_name": supplier_name,
                "supplier_key": supplier_match_key(supplier_name),
                "opening_debit": _money(ws.cell(r, 15).value),
                "opening_credit": _money(ws.cell(r, 16).value),
                "period_debit": _money(ws.cell(r, 18).value),
                "period_credit": _money(ws.cell(r, 20).value),
                "ending_debit": _money(ws.cell(r, 22).value),
                "ending_credit": ending_credit,
                "payable_amount": ending_credit,
                "row_order": len(rows) + 1,
            }
        )

    return rows


def parse_import(path: Path, source_type: str) -> list[dict[str, Any]]:
    if source_type == "DNTT":
        return parse_dntt_xlsx(path)
    if source_type == "VON_TU_CO":
        return parse_von_tu_co_xlsx(path)
    if source_type == "THCN_PAYABLE":
        return parse_thcn_payable_xlsx(path)
    raise ValueError("source_type không hợp lệ")


def money_fmt(value: Any) -> str:
    n = _money(value)
    return f"{n:,.0f}".replace(",", ".")


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_merge(ws, cell_range: str, value: Any, bold: bool = False, size: int = 11, center: bool = False) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.font = Font(name="Times New Roman", size=size, bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)


def _doc_check_amount(doc) -> float:
    if doc["ksnb_check_type"] == "DIFFERENT":
        return _money(doc["ksnb_checked_amount"])
    return _money(doc["document_amount"])


def _doc_payment_note(doc) -> str:
    methods = []
    if doc["payment_transfer"]:
        methods.append("Chuyển khoản")
    if doc["payment_cash"]:
        methods.append("Tiền mặt")
    note = (doc["ksnb_note"] or "").strip()
    if methods and note:
        return "; ".join(methods) + "; " + note
    if methods:
        return "; ".join(methods)
    return note


def export_voucher_xlsx(voucher_id: int) -> Path:
    with get_conn() as conn:
        voucher = conn.execute("SELECT v.*, u.full_name AS creator_name, u.position_title AS creator_position FROM vouchers v JOIN users u ON u.id = v.created_by WHERE v.id = ?", (voucher_id,)).fetchone()
        review_items = conn.execute("SELECT * FROM voucher_review_items WHERE voucher_id = ? ORDER BY item_order", (voucher_id,)).fetchall()
        docs = conn.execute("SELECT * FROM voucher_documents WHERE voucher_id = ? ORDER BY row_order", (voucher_id,)).fetchall()
        signatures = conn.execute("SELECT * FROM voucher_signatures WHERE voucher_id = ? ORDER BY id", (voucher_id,)).fetchall()
    if not voucher:
        raise ValueError("Không tìm thấy Phiếu")

    doc_total = sum(_money(d["document_amount"]) for d in docs)
    ksnb_total = sum(_doc_check_amount(d) for d in docs)
    display_total = doc_total if docs else _money(voucher["total_amount"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Phiếu KSNB"
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left = 0.45
    ws.page_margins.right = 0.45
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    widths = {"A": 5, "B": 24, "C": 36, "D": 16, "E": 18, "F": 20, "G": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    row = 1
    _set_merge(ws, "A1:C1", "CÔNG TY TNHH BỆNH VIỆN", bold=True, center=True)
    _set_merge(ws, "E1:G1", "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM", bold=True, center=True)
    _set_merge(ws, "A2:C2", "HÙNG VƯƠNG GIA LAI", bold=True, center=True)
    _set_merge(ws, "E2:G2", "Độc lập - Tự do - Hạnh phúc", bold=True, center=True)
    _set_merge(ws, "A3:C3", "BAN KIỂM SOÁT NỘI BỘ", bold=True, center=True)
    _set_merge(ws, "E3:G3", f"Gia Lai, ngày {datetime.now().day:02d} tháng {datetime.now().month:02d} năm {datetime.now().year}", center=True)
    row = 5
    _set_merge(ws, f"A{row}:G{row}", "PHIẾU KIỂM SOÁT TRƯỚC HỒ SƠ THU, CHI QUAN TRỌNG", bold=True, size=14, center=True)

    row += 2
    _set_merge(ws, f"A{row}:G{row}", "I. THÔNG TIN CHUNG", bold=True)
    info = [
        ("1. Tên hồ sơ/khoản thu, chi", voucher["title"]),
        ("2. Loại hồ sơ", "☑ Thu quan trọng" if voucher["hồ_so_type"] == "THU" else "☑ Chi quan trọng"),
        ("3. Số tiền", money_fmt(display_total)),
        ("4. Đơn vị lập hồ sơ", voucher["submitting_unit"] or ""),
        ("5. Người giao hồ sơ", voucher["sender_name"] or ""),
        ("6. Ngày tiếp nhận hồ sơ tại Ban Kiểm soát nội bộ", voucher["received_at"] or ""),
        ("7. Hình thức trình phê duyệt", voucher["approval_target"] or "Trình Hội đồng thành viên"),
    ]
    for label, value in info:
        row += 1
        _set_merge(ws, f"A{row}:C{row}", label + ":", bold=True)
        _set_merge(ws, f"D{row}:G{row}", value)

    row += 2
    _set_merge(ws, f"A{row}:G{row}", "II. NỘI DUNG KIỂM SOÁT TRƯỚC", bold=True)
    row += 1
    ws[f"A{row}"] = "STT"
    ws[f"B{row}"] = "Nội dung rà soát, kiểm soát"
    ws[f"G{row}"] = "Ghi chú"
    ws.merge_cells(f"B{row}:F{row}")
    for c in ["A", "B", "G"]:
        ws[f"{c}{row}"].font = Font(name="Times New Roman", bold=True)
        ws[f"{c}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"{c}{row}"].border = _thin_border()
    for idx, item in enumerate(review_items, start=1):
        row += 1
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = item["content"]
        ws[f"G{row}"] = item["note"] or ""
        ws.merge_cells(f"B{row}:F{row}")
        for c in ["A", "B", "G"]:
            ws[f"{c}{row}"].alignment = Alignment(vertical="center", wrap_text=True)
            ws[f"{c}{row}"].border = _thin_border()

    row += 2
    _set_merge(ws, f"A{row}:G{row}", "III. DANH MỤC HỒ SƠ, TÀI LIỆU NHẬN KIỂM SOÁT", bold=True)
    row += 1
    doc_headers = ["STT", "Nhà cung cấp", "Nội dung", "Số tiền", "Ban KSNB KT", "Ghi chú"]
    for col, text in zip(["A", "B", "C", "D", "E", "F"], doc_headers):
        ws[f"{col}{row}"] = text
        ws[f"{col}{row}"].font = Font(name="Times New Roman", bold=True)
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"{col}{row}"].border = _thin_border()
    ws.merge_cells(f"F{row}:G{row}")
    for idx, doc in enumerate(docs, start=1):
        row += 1
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = doc["supplier"] or ""
        ws[f"C{row}"] = doc["content"] or ""
        ws[f"D{row}"] = _money(doc["document_amount"])
        ws[f"D{row}"].number_format = '#,##0'
        result = CHECK_MARK if doc["ksnb_check_type"] == "MATCH" else money_fmt(doc["ksnb_checked_amount"])
        ws[f"E{row}"] = result
        ws[f"F{row}"] = _doc_payment_note(doc)
        ws.merge_cells(f"F{row}:G{row}")
        for c in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{c}{row}"].alignment = Alignment(horizontal="center" if c in ["A", "D", "E"] else "left", vertical="center", wrap_text=True)
            ws[f"{c}{row}"].border = _thin_border()
    if docs:
        row += 1
        _set_merge(ws, f"A{row}:C{row}", "Tổng cộng", bold=True, center=True)
        ws[f"D{row}"] = doc_total
        ws[f"E{row}"] = ksnb_total
        ws[f"D{row}"].number_format = '#,##0'
        ws[f"E{row}"].number_format = '#,##0'
        ws.merge_cells(f"F{row}:G{row}")
        for c in ["A", "D", "E", "F"]:
            ws[f"{c}{row}"].font = Font(name="Times New Roman", bold=True)
            ws[f"{c}{row}"].alignment = Alignment(horizontal="center" if c != "D" and c != "E" else "right", vertical="center", wrap_text=True)
            ws[f"{c}{row}"].border = _thin_border()

    row += 2
    _set_merge(ws, f"A{row}:G{row}", "IV. KẾT QUẢ KIỂM SOÁT TRƯỚC", bold=True)
    row += 1
    _set_merge(ws, f"A{row}:G{row}", voucher["section_iv_result"] or "")
    if voucher["section_iv_note"]:
        row += 1
        _set_merge(ws, f"A{row}:G{row}", voucher["section_iv_note"] or "")

    section_index = 5
    if (voucher["section_v_text"] or "").strip():
        row += 2
        roman = {5: "V", 6: "VI", 7: "VII"}.get(section_index, str(section_index))
        _set_merge(ws, f"A{row}:G{row}", f"{roman}. KIẾN NGHỊ CỦA BAN KIỂM SOÁT NỘI BỘ", bold=True)
        row += 1
        _set_merge(ws, f"A{row}:G{row}", voucher["section_v_text"].strip())
        section_index += 1
    if (voucher["section_vi_text"] or "").strip():
        row += 2
        roman = {5: "V", 6: "VI", 7: "VII"}.get(section_index, str(section_index))
        _set_merge(ws, f"A{row}:G{row}", f"{roman}. Ý KIẾN CỦA ĐƠN VỊ LẬP HỒ SƠ", bold=True)
        row += 1
        _set_merge(ws, f"A{row}:G{row}", voucher["section_vi_text"].strip())
        section_index += 1

    row += 2
    roman = {5: "V", 6: "VI", 7: "VII"}.get(section_index, str(section_index))
    _set_merge(ws, f"A{row}:G{row}", f"{roman}. KẾT LUẬN CỦA BAN KIỂM SOÁT NỘI BỘ", bold=True)
    row += 1
    _set_merge(ws, f"A{row}:G{row}", CONCLUSION_TEXT)

    row += 3
    _set_merge(ws, f"A{row}:C{row}", "NGƯỜI KIỂM SOÁT", bold=True, center=True)
    _set_merge(ws, f"E{row}:G{row}", "TRƯỞNG BAN KIỂM SOÁT NỘI BỘ", bold=True, center=True)
    row += 1
    creator_sig = next((s for s in signatures if s["signer_role"] == "NGUOI_KIEM_SOAT"), None)
    head_sig = next((s for s in signatures if s["signer_role"] == "TRUONG_BAN"), None)
    creator_text = creator_sig["signature_text"] if creator_sig else ""
    if voucher["route_mode"] == "DIRECT_BOARD":
        head_text = HEAD_DELEGATION_TEXT
    else:
        head_text = head_sig["signature_text"] if head_sig else ""
    _set_merge(ws, f"A{row}:C{row+2}", creator_text, center=True)
    _set_merge(ws, f"E{row}:G{row+2}", head_text, center=True)

    for r in range(1, row + 4):
        ws.row_dimensions[r].height = 24
    ws.freeze_panes = "A1"
    output = EXPORT_DIR / f"phieu_ksnb_{voucher_id}.xlsx"
    wb.save(output)
    return output


def convert_xlsx_to_pdf(xlsx_path: Path) -> Path | None:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(EXPORT_DIR), str(xlsx_path)],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    pdf_path = EXPORT_DIR / (xlsx_path.stem + ".pdf")
    return pdf_path if pdf_path.exists() else None
