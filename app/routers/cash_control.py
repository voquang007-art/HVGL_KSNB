from __future__ import annotations

import re
import unicodedata
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from ..database import BOARD_ROLES, HEAD_ROLES, IMPORT_DIR, can_create_voucher, get_conn
from ..excel_service import REVIEW_ITEM_MASTER
from ..upload_security import (
    BUSINESS_IMPORT_EXTENSIONS,
    BUSINESS_IMPORT_MAX_UPLOAD_MB,
    CURRENT_EXCEL_PARSER_EXTENSIONS,
    UploadValidationError,
    safe_original_filename,
    save_upload_file_chunked,
)

templates = Jinja2Templates(directory="app/templates")
router = APIRouter(prefix="/cash-control")

EMPLOYEE_LINE_RE = re.compile(r"^\s*(\d+)\s*[-–]\s*(.+?)\s*$")

CASH_CONTROL_REVENUE_ITEMS = [
    "Thu tiền khám chữa bệnh",
    "Thu tiền Căn tin",
    "Thu tiền tiêm chủng",
    "Thu tiền Da Liễu",
    "Thu tiền Nhà xe",
    "Thu tiền Siêu Thị",
    "Thu tiền Nhà Thuốc",
    "Thu tiền Hồ Sơ Bệnh Án",
    "Thu tiền khác",
]

CASH_CONTROL_EXPENSE_ITEMS = [
    "Chi hoàn tiền Khám chữa bệnh",
    "Chi khác",
]

CASH_CONTROL_STATUS_LABELS = {
    "DRAFT": "Nháp",
    "SUBMITTED_TO_HEAD": "Đã trình Trưởng/Phó Ban KSNB",
    "SUBMITTED_TO_BOARD": "Đã trình HĐTV",
    "BOARD_VIEWED": "HĐTV đã xem",
    "BOARD_SAVED": "HĐTV đã lưu",
    "NO_SIGNATURE_INTERNAL": "Không ký số - lưu nội bộ",
    "RETURNED": "Đã trả lại",
}

RECONCILE_REVENUE_ITEMS = [
    "Thu tiền khám chữa bệnh",
    "Thu tiền Căn tin",
    "Thu tiền tiêm chủng",
    "Thu tiền Da Liễu",
    "Thu tiền Nhà xe",
    "Thu tiền Siêu Thị",
    "Thu tiền Nhà Thuốc",
    "Thu tiền Hồ Sơ Bệnh Án",
    "Thu tiền khác",
]

ACCOUNTING_REVENUE_ALIASES = {
    "Thu tiền khám chữa bệnh": ["thuc thu tien mat", "tien kham chua benh", "kham chua benh"],
    "Thu tiền Căn tin": ["can tin", "cantin", "can-tin"],
    "Thu tiền tiêm chủng": ["tiem chung"],
    "Thu tiền Da Liễu": ["da lieu"],
    "Thu tiền Nhà xe": ["nha xe", "giu xe"],
    "Thu tiền Siêu Thị": ["sieu thi"],
    "Thu tiền Nhà Thuốc": ["nha thuoc"],
    "Thu tiền Hồ Sơ Bệnh Án": ["ho so benh an", "hsba", "giay chung nhan thuong tich"],
    "Thu tiền khác": ["thu ung", "hoan ung", "thu khac", "ung nhan vien"],
}

VN_TZ = timezone(timedelta(hours=7))

def _save_cash_control_import_upload(upload: UploadFile, stored_path: Path) -> str:
    try:
        _size_bytes, _sha256_value, original_name = save_upload_file_chunked(
            upload,
            stored_path,
            allowed_extensions=BUSINESS_IMPORT_EXTENSIONS,
            max_size_mb=BUSINESS_IMPORT_MAX_UPLOAD_MB,
            parser_supported_extensions=CURRENT_EXCEL_PARSER_EXTENSIONS,
        )
    except UploadValidationError as ex:
        raise ValueError(str(ex)) from ex
    return original_name
    
def format_vn_sqlite_datetime(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return text

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    vn_time = parsed.astimezone(VN_TZ)
    return vn_time.strftime("%d/%m/%Y %H:%M:%S")
    

def format_vn_control_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return text

templates.env.globals["format_vn_control_date"] = format_vn_control_date    

def query_int(request: Request, name: str) -> int | None:
    value = request.query_params.get(name)
    if value is None or str(value).strip() == "":
        return None

    try:
        return int(value)
    except ValueError:
        return None


def query_text(request: Request, name: str) -> str:
    return str(request.query_params.get(name) or "").strip()


def sqlite_datetime_parts(value: Any) -> tuple[str, str, str]:
    text = str(value or "").strip()
    if not text:
        return "", "", ""

    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return "", "", ""

    return f"{parsed.year:04d}", f"{parsed.month:02d}", f"{parsed.day:02d}"


def filter_import_batches(
    rows: list[Any],
    *,
    selected_year: str = "",
    selected_month: str = "",
    selected_day: str = "",
) -> list[Any]:
    filtered_rows: list[Any] = []
    for row in rows:
        year, month, day = sqlite_datetime_parts(row["created_at"])
        if selected_year and year != selected_year:
            continue
        if selected_month and month != selected_month.zfill(2):
            continue
        if selected_day and day != selected_day.zfill(2):
            continue
        filtered_rows.append(row)
    return filtered_rows


def import_filter_choices(rows: list[Any]) -> dict[str, list[str]]:
    years: set[str] = set()
    months: set[str] = set()
    days: set[str] = set()
    for row in rows:
        year, month, day = sqlite_datetime_parts(row["created_at"])
        if year:
            years.add(year)
        if month:
            months.add(month)
        if day:
            days.add(day)
    return {
        "years": sorted(years, reverse=True),
        "months": sorted(months),
        "days": sorted(days),
    }

    
def require_login(request: Request):
    if not request.state.user:
        return RedirectResponse("/login", status_code=303)
    return None


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    return " ".join(text.split())


def classify_cash_receipt(counterparty_credit_code: Any, description: Any) -> str:
    code = str(counterparty_credit_code or "").strip().upper()
    desc = normalize_text(description)

    if code in {"KHACHKCB", "KSKSK", "KLKSK", "THAMMY", "QUAYKINH"}:
        return "Thu tiền khám chữa bệnh"
    if code == "CANTIN":
        return "Thu tiền Căn tin"
    if code == "TIEMCHUNG":
        return "Thu tiền tiêm chủng"
    if code == "DALIEU":
        return "Thu tiền Da Liễu"
    if code == "GUIXE":
        return "Thu tiền Nhà xe"
    if code == "SIEUTHI":
        return "Thu tiền Siêu Thị"
    if code == "NHATHUOC":
        return "Thu tiền Nhà Thuốc"
    if code == "KHACHKHAC" and (
        "ho so benh an" in desc
        or "giay chung nhan thuong tich" in desc
        or "sao benh an" in desc
        or "trich luc ho so" in desc
    ):
        return "Thu tiền Hồ Sơ Bệnh Án"
    return "Thu tiền khác"


def classify_cash_payment(counterparty_debit_code: Any, description: Any) -> str:
    code = str(counterparty_debit_code or "").strip().upper()

    if code == "KHACHKCB":
        return "Chi hoàn tiền Khám chữa bệnh"
    return "Chi khác"


def find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        current_values = [normalize_text(cell.value) for cell in ws[row_idx]]
        next_values = [normalize_text(cell.value) for cell in ws[row_idx + 1]] if row_idx < ws.max_row else []
        joined = " | ".join(current_values)

        if "loai" in joined and "dien giai" in joined and ("so tien" in joined or "vnd" in joined):
            columns: dict[str, int] = {}
            last_main_header = ""

            for idx, value in enumerate(current_values):
                if value:
                    last_main_header = value

                sub_value = next_values[idx] if idx < len(next_values) else ""
                combined = " ".join(part for part in [last_main_header, sub_value] if part).strip()

                if combined:
                    columns[combined] = idx
                if value:
                    columns[value] = idx

            return row_idx + 1, columns

    return 1, {normalize_text(cell.value): idx for idx, cell in enumerate(ws[1]) if cell.value}


def find_column(columns: dict[str, int], candidates: list[str]) -> int | None:
    normalized_candidates = [normalize_text(candidate) for candidate in candidates]
    for header, index in columns.items():
        for candidate in normalized_candidates:
            if candidate and candidate in header:
                return index
    return None


def get_row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def money_value(value: Any) -> float:
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = text.replace(" ", "").replace("\u00a0", "")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]

    text = text.replace("đ", "").replace("Đ", "").replace("VND", "").replace("vnd", "")
    text = text.strip()

    if not text:
        return 0.0

    if "," in text and "." in text:
        # Định dạng Việt Nam: 1.234.567,89
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        # Định dạng có dấu phẩy thập phân hoặc phân tách phần nghìn.
        parts = text.split(",")
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) in (1, 2):
            text = parts[0].replace(".", "").replace(",", "") + "." + parts[1]
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) in (1, 2):
            # Trường hợp Excel/openpyxl trả về chuỗi số dạng 123456.0 hoặc 123456.00
            # thì dấu chấm là dấu thập phân, không được xóa.
            text = parts[0] + "." + parts[1]
        else:
            # Trường hợp 1.234.567 là phân tách phần nghìn.
            text = text.replace(".", "")

    try:
        amount = float(text)
    except ValueError:
        return 0.0

    return -amount if negative else amount


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def is_detail_row(stt_value: Any, ma_kcb_value: Any) -> bool:
    if is_blank(ma_kcb_value):
        return False
    if isinstance(stt_value, int):
        return True
    if isinstance(stt_value, float) and stt_value.is_integer():
        return True
    return str(stt_value or "").strip().isdigit()


def parse_employee_line(value: Any) -> tuple[str, str] | None:
    text = str(value or "").strip()
    match = EMPLOYEE_LINE_RE.match(text)
    if not match:
        return None
    return match.group(1).strip(), match.group(2).strip()


def parse_revenue_excel(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    current_employee_code = ""
    current_employee_name = ""
    summaries: dict[str, dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=1, values_only=True):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None

        col_a_text = str(col_a or "").strip()
        if col_a_text.startswith("Tổng cộng"):
            break

        employee_info = parse_employee_line(col_a)
        if employee_info and is_blank(col_b):
            current_employee_code, current_employee_name = employee_info
            key = current_employee_code
            summaries.setdefault(
                key,
                {
                    "employee_code": current_employee_code,
                    "employee_name": current_employee_name,
                    "patient_codes": set(),
                    "patient_count": 0,
                    "txn_count": 0,
                    "tien_thu": 0.0,
                    "tam_ung": 0.0,
                    "tien_chi": 0.0,
                    "mien_giam": 0.0,
                    "huy_phieu": 0.0,
                    "thuc_thu": 0.0,
                    "tien_mat": 0.0,
                    "chuyen_khoan": 0.0,
                    "qr": 0.0,
                    "pos": 0.0,
                    "details": [],
                },
            )
            continue

        if not is_detail_row(col_a, col_b):
            continue

        employee_code = current_employee_code or "UNKNOWN"
        employee_name = current_employee_name or "Chưa xác định nhân viên"
        item = summaries.setdefault(
            employee_code,
            {
                "employee_code": employee_code,
                "employee_name": employee_name,
                "patient_codes": set(),
                "patient_count": 0,
                "txn_count": 0,
                "tien_thu": 0.0,
                "tam_ung": 0.0,
                "tien_chi": 0.0,
                "mien_giam": 0.0,
                "huy_phieu": 0.0,
                "thuc_thu": 0.0,
                "tien_mat": 0.0,
                "chuyen_khoan": 0.0,
                "qr": 0.0,
                "pos": 0.0,
                "details": [],
            },
        )

        item["txn_count"] += 1
        patient_code = str(col_b or "").strip()
        if patient_code:
            item["patient_codes"].add(patient_code)

        # Cột theo file báo cáo:
        # L: Tiền thu, M: Tạm ứng, N: Tiền chi, O: Miễn giảm,
        # Q: Hủy phiếu, S: Thực thu, T: Tiền mặt, U: Chuyển khoản, V: QR, W: POS
        item["tien_thu"] += money_value(row[11] if len(row) > 11 else None)
        item["tam_ung"] += money_value(row[12] if len(row) > 12 else None)
        item["tien_chi"] += money_value(row[13] if len(row) > 13 else None)
        item["mien_giam"] += money_value(row[14] if len(row) > 14 else None)
        item["huy_phieu"] += money_value(row[16] if len(row) > 16 else None)
        item["thuc_thu"] += money_value(row[18] if len(row) > 18 else None)
        item["tien_mat"] += money_value(row[19] if len(row) > 19 else None)
        item["chuyen_khoan"] += money_value(row[20] if len(row) > 20 else None)
        item["qr"] += money_value(row[21] if len(row) > 21 else None)
        item["pos"] += money_value(row[22] if len(row) > 22 else None)
        item["details"].append(
            {
                "employee_code": employee_code,
                "employee_name": employee_name,
                "patient_code": patient_code,
                "patient_name": str(row[2] if len(row) > 2 and row[2] is not None else "").strip(),
                "tien_thu": money_value(row[11] if len(row) > 11 else None),
                "tam_ung": money_value(row[12] if len(row) > 12 else None),
                "tien_chi": money_value(row[13] if len(row) > 13 else None),
                "huy_phieu": money_value(row[16] if len(row) > 16 else None),
                "thuc_thu": money_value(row[18] if len(row) > 18 else None),
                "tien_mat": money_value(row[19] if len(row) > 19 else None),
                "chuyen_khoan": money_value(row[20] if len(row) > 20 else None),
                "qr": money_value(row[21] if len(row) > 21 else None),
                "pos": money_value(row[22] if len(row) > 22 else None),
            }
        )

    for item in summaries.values():
        item["patient_count"] = len(item.get("patient_codes", set()))
        item.pop("patient_codes", None)

    return summaries


def parse_cashbook_excel(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row, columns = find_header_row_and_columns(ws)

    type_col = find_column(columns, ["loại c.từ", "loại chứng từ", "loại"])
    description_col = find_column(columns, ["diễn giải"])
    debit_code_col = find_column(columns, ["mã đối tượng nợ"])
    credit_code_col = find_column(columns, ["mã đối tượng có"])
    debit_amount_col = find_column(columns, ["số tiền vnd nợ"])
    credit_amount_col = find_column(columns, ["số tiền vnd có"])
    voucher_no_col = find_column(columns, ["chứng từ số", "số chứng từ", "số ct", "số phiếu"])
    voucher_date_col = find_column(columns, ["chứng từ ngày", "ngày chứng từ", "ngày"])

    receipt_summary = {
        item: {"item_name": item, "voucher_numbers": set(), "voucher_count": 0, "amount": 0.0}
        for item in CASH_CONTROL_REVENUE_ITEMS
    }
    expense_summary = {
        item: {"item_name": item, "voucher_numbers": set(), "voucher_count": 0, "amount": 0.0}
        for item in CASH_CONTROL_EXPENSE_ITEMS
    }

    receipt_numbers: set[str] = set()
    expense_numbers: set[str] = set()
    detail_entries: list[dict[str, Any]] = []

    if type_col is None or description_col is None or debit_code_col is None or credit_code_col is None:
        return {
            "receipt_rows": list(receipt_summary.values()),
            "expense_rows": list(expense_summary.values()),
            "receipt_count": 0,
            "expense_count": 0,
            "refund_kcb_count": 0,
            "receipt_total": 0.0,
            "expense_total": 0.0,
        }

    if debit_amount_col is None or credit_amount_col is None:
        return {
            "receipt_rows": list(receipt_summary.values()),
            "expense_rows": list(expense_summary.values()),
            "receipt_count": 0,
            "expense_count": 0,
            "refund_kcb_count": 0,
            "receipt_total": 0.0,
            "expense_total": 0.0,
        }

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        doc_type = str(get_row_value(row, type_col) or "").strip().upper()
        if doc_type not in {"PT", "PC"}:
            continue

        description = get_row_value(row, description_col)
        debit_code = get_row_value(row, debit_code_col)
        credit_code = get_row_value(row, credit_code_col)
        debit_amount = money_value(get_row_value(row, debit_amount_col))
        credit_amount = money_value(get_row_value(row, credit_amount_col))
        voucher_no = str(get_row_value(row, voucher_no_col) or "").strip()
        if not voucher_no:
            voucher_no = f"ROW-{row_idx}"

        if doc_type == "PT" and debit_amount > 0:
            item_name = classify_cash_receipt(credit_code, description)
            receipt_numbers.add(voucher_no)
            receipt_summary.setdefault(
                item_name,
                {"item_name": item_name, "voucher_numbers": set(), "voucher_count": 0, "amount": 0.0},
            )
            receipt_summary[item_name]["voucher_numbers"].add(voucher_no)
            receipt_summary[item_name]["amount"] += debit_amount
            detail_entries.append(
                {
                    "doc_type": "PT",
                    "voucher_no": voucher_no,
                    "voucher_date": str(get_row_value(row, voucher_date_col) or "").strip(),
                    "item_name": item_name,
                    "debit_code": str(debit_code or "").strip(),
                    "credit_code": str(credit_code or "").strip(),
                    "description": str(description or "").strip(),
                    "amount": debit_amount,
                }
            )

        if doc_type == "PC" and credit_amount > 0:
            item_name = classify_cash_payment(debit_code, description)
            expense_numbers.add(voucher_no)
            expense_summary.setdefault(
                item_name,
                {"item_name": item_name, "voucher_numbers": set(), "voucher_count": 0, "amount": 0.0},
            )
            expense_summary[item_name]["voucher_numbers"].add(voucher_no)
            expense_summary[item_name]["amount"] += credit_amount
            detail_entries.append(
                {
                    "doc_type": "PC",
                    "voucher_no": voucher_no,
                    "voucher_date": str(get_row_value(row, voucher_date_col) or "").strip(),
                    "item_name": item_name,
                    "debit_code": str(debit_code or "").strip(),
                    "credit_code": str(credit_code or "").strip(),
                    "description": str(description or "").strip(),
                    "amount": credit_amount,
                }
            )

    for item in receipt_summary.values():
        item["voucher_count"] = len(item.get("voucher_numbers", set()))
        item.pop("voucher_numbers", None)

    for item in expense_summary.values():
        item["voucher_count"] = len(item.get("voucher_numbers", set()))
        item.pop("voucher_numbers", None)

    receipt_rows = [receipt_summary[item] for item in CASH_CONTROL_REVENUE_ITEMS]
    expense_rows = [expense_summary[item] for item in CASH_CONTROL_EXPENSE_ITEMS]

    receipt_total = sum(money_value(item["amount"]) for item in receipt_rows)
    expense_total = sum(money_value(item["amount"]) for item in expense_rows)
    refund_kcb_count = int(expense_summary["Chi hoàn tiền Khám chữa bệnh"]["voucher_count"])

    return {
        "receipt_rows": receipt_rows,
        "expense_rows": expense_rows,
        "receipt_count": len(receipt_numbers),
        "expense_count": len(expense_numbers),
        "refund_kcb_count": refund_kcb_count,
        "receipt_total": receipt_total,
        "expense_total": expense_total,
        "detail_entries": detail_entries,
    }

def classify_accounting_revenue_item(label: Any) -> str | None:
    text = normalize_text(label)
    compact_text = text.replace(" ", "")

    for item_name, aliases in ACCOUNTING_REVENUE_ALIASES.items():
        for alias in aliases:
            normalized_alias = normalize_text(alias)
            compact_alias = normalized_alias.replace(" ", "")
            if normalized_alias in text or compact_alias in compact_text:
                return item_name

    return None


def parse_accounting_excel(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    employee_rows: list[dict[str, Any]] = []
    summary_map = {item: {"item_name": item, "amount": 0.0, "note": ""} for item in RECONCILE_REVENUE_ITEMS}
    refund_kcb_amount = 0.0
    accounting_non_cash_diff_total = 0.0
    has_total_kcb_cash_row = False

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        text_values = [str(v).strip() for v in values if v is not None and str(v).strip()]
        row_text = " | ".join(text_values)
        norm_row_text = normalize_text(row_text)

        if not row_text:
            continue

        first_text = str(values[0] if len(values) > 0 and values[0] is not None else "").strip()
        norm_first_text = normalize_text(first_text)

        employee_info = parse_employee_line(first_text)
        if employee_info:
            employee_code, employee_name = employee_info
            accounting_bank_ml_amount = money_value(values[2] if len(values) > 2 else None)
            accounting_qr_amount = money_value(values[3] if len(values) > 3 else None)
            accounting_pos_amount = money_value(values[4] if len(values) > 4 else None)
            accounting_report_non_cash_amount = money_value(values[5] if len(values) > 5 else None)
            accounting_non_cash_amount = (
                accounting_bank_ml_amount
                + accounting_qr_amount
                + accounting_pos_amount
            )
            accounting_non_cash_diff_amount = (
                accounting_non_cash_amount
                - accounting_report_non_cash_amount
            )
            accounting_cash_amount = money_value(values[6] if len(values) > 6 else None)

            employee_rows.append(
                {
                    "employee_code": employee_code,
                    "employee_name": employee_name,
                    "accounting_cash_amount": accounting_cash_amount,
                    "accounting_bank_ml_amount": accounting_bank_ml_amount,
                    "accounting_qr_amount": accounting_qr_amount,
                    "accounting_pos_amount": accounting_pos_amount,
                    "accounting_non_cash_amount": accounting_non_cash_amount,
                    "accounting_report_non_cash_amount": accounting_report_non_cash_amount,
                    "accounting_non_cash_diff_amount": accounting_non_cash_diff_amount,
                    "note": str(values[8] if len(values) > 8 and values[8] is not None else "").strip(),
                }
            )
            continue

        if norm_first_text in {"tong", "tong cong"}:
            amount = money_value(values[6] if len(values) > 6 else None)
            accounting_non_cash_diff_total = money_value(values[7] if len(values) > 7 else None)
            summary_map["Thu tiền khám chữa bệnh"]["amount"] += amount
            summary_map["Thu tiền khám chữa bệnh"]["note"] = first_text
            has_total_kcb_cash_row = True
            continue

        if norm_first_text.startswith("tong thu"):
            tong_thu_diff_amount = money_value(values[7] if len(values) > 7 else None)
            if tong_thu_diff_amount or accounting_non_cash_diff_total == 0:
                accounting_non_cash_diff_total = tong_thu_diff_amount
            continue

        matched_item_name = classify_accounting_revenue_item(row_text)
        if matched_item_name:
            amount = money_value(values[6] if len(values) > 6 else None)
            summary_map[matched_item_name]["amount"] += amount
            summary_map[matched_item_name]["note"] = row_text
            continue

        if first_text and not EMPLOYEE_LINE_RE.match(first_text):
            accounting_cash_amount = money_value(values[6] if len(values) > 6 else None)
            if accounting_cash_amount > 0:
                accounting_bank_ml_amount = money_value(values[2] if len(values) > 2 else None)
                accounting_qr_amount = money_value(values[3] if len(values) > 3 else None)
                accounting_pos_amount = money_value(values[4] if len(values) > 4 else None)
                accounting_report_non_cash_amount = money_value(values[5] if len(values) > 5 else None)

                employee_rows.append(
                    {
                        "employee_code": "",
                        "employee_name": first_text,
                        "accounting_cash_amount": accounting_cash_amount,
                        "accounting_bank_ml_amount": accounting_bank_ml_amount,
                        "accounting_qr_amount": accounting_qr_amount,
                        "accounting_pos_amount": accounting_pos_amount,
                        "accounting_non_cash_amount": (
                            accounting_bank_ml_amount
                            + accounting_qr_amount
                            + accounting_pos_amount
                        ),
                        "accounting_report_non_cash_amount": accounting_report_non_cash_amount,
                        "accounting_non_cash_diff_amount": money_value(values[7] if len(values) > 7 else None),
                        "note": str(values[8] if len(values) > 8 and values[8] is not None else "").strip(),
                    }
                )
                continue

        matched_item_name = classify_accounting_revenue_item(row_text)
        if matched_item_name:
            amount = money_value(values[6] if len(values) > 6 else None)
            summary_map[matched_item_name]["amount"] += amount
            summary_map[matched_item_name]["note"] = row_text
            continue

        if "hoan" in norm_row_text and "kham" in norm_row_text and "chua benh" in norm_row_text:
            numeric_values = [money_value(v) for v in values if money_value(v) != 0]
            refund_kcb_amount += abs(numeric_values[-1]) if numeric_values else 0.0

    if not has_total_kcb_cash_row:
        summary_map["Thu tiền khám chữa bệnh"]["amount"] = sum(
            money_value(row.get("accounting_cash_amount")) for row in employee_rows
        )
        summary_map["Thu tiền khám chữa bệnh"]["note"] = "Tổng tiền mặt theo từng nhân viên"

    return {
        "employee_rows": employee_rows,
        "summary_rows": list(summary_map.values()),
        "refund_kcb_amount": refund_kcb_amount,
        "accounting_non_cash_diff_total": accounting_non_cash_diff_total,
    }
    
    
def ensure_cash_control_tables() -> None:
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS cash_control_cashbook_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                original_filename TEXT,
                created_by INTEGER NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS cash_control_cashbook_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                summary_type TEXT NOT NULL,
                item_name TEXT NOT NULL,
                voucher_count INTEGER NOT NULL DEFAULT 0,
                amount REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(batch_id) REFERENCES cash_control_cashbook_batches(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_cashbook_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                doc_type TEXT NOT NULL,
                voucher_no TEXT,
                voucher_date TEXT,
                item_name TEXT NOT NULL,
                debit_code TEXT,
                credit_code TEXT,
                description TEXT,
                amount REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(batch_id) REFERENCES cash_control_cashbook_batches(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_his_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                employee_code TEXT,
                employee_name TEXT,
                patient_code TEXT,
                patient_name TEXT,
                tien_thu REAL NOT NULL DEFAULT 0,
                tam_ung REAL NOT NULL DEFAULT 0,
                tien_chi REAL NOT NULL DEFAULT 0,
                huy_phieu REAL NOT NULL DEFAULT 0,
                thuc_thu REAL NOT NULL DEFAULT 0,
                tien_mat REAL NOT NULL DEFAULT 0,
                chuyen_khoan REAL NOT NULL DEFAULT 0,
                qr REAL NOT NULL DEFAULT 0,
                pos REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(batch_id) REFERENCES cash_control_batches(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_accounting_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                original_filename TEXT,
                created_by INTEGER NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                accounting_non_cash_diff_total REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS cash_control_accounting_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                item_name TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                note TEXT,
                FOREIGN KEY(batch_id) REFERENCES cash_control_accounting_batches(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_accounting_employee_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                employee_code TEXT,
                employee_name TEXT,
                his_cash_amount REAL NOT NULL DEFAULT 0,
                accounting_cash_amount REAL NOT NULL DEFAULT 0,
                accounting_bank_ml_amount REAL NOT NULL DEFAULT 0,
                accounting_qr_amount REAL NOT NULL DEFAULT 0,
                accounting_pos_amount REAL NOT NULL DEFAULT 0,
                accounting_non_cash_amount REAL NOT NULL DEFAULT 0,
                accounting_report_non_cash_amount REAL NOT NULL DEFAULT 0,
                accounting_non_cash_diff_amount REAL NOT NULL DEFAULT 0,
                diff_amount REAL NOT NULL DEFAULT 0,
                note TEXT,
                FOREIGN KEY(batch_id) REFERENCES cash_control_accounting_batches(id) ON DELETE CASCADE
            );
            
            CREATE TABLE IF NOT EXISTS cash_control_vouchers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT,
                title TEXT NOT NULL,
                voucher_date TEXT,
                sign_mode TEXT NOT NULL DEFAULT 'INTERNAL_DIGITAL',
                route_mode TEXT NOT NULL DEFAULT 'THROUGH_HEAD',
                head_user_id INTEGER,
                board_user_id INTEGER,
                current_handler INTEGER,
                his_batch_id INTEGER,
                cashbook_batch_id INTEGER,
                accounting_batch_id INTEGER,
                receipt_count INTEGER NOT NULL DEFAULT 0,
                expense_count INTEGER NOT NULL DEFAULT 0,
                refund_kcb_count INTEGER NOT NULL DEFAULT 0,
                revenue_document_total REAL NOT NULL DEFAULT 0,
                revenue_ksnb_total REAL NOT NULL DEFAULT 0,
                revenue_diff_total REAL NOT NULL DEFAULT 0,
                expense_document_total REAL NOT NULL DEFAULT 0,
                expense_ksnb_total REAL NOT NULL DEFAULT 0,
                expense_diff_total REAL NOT NULL DEFAULT 0,
                section_iv_result TEXT,
                section_iv_note TEXT,
                section_v_text TEXT,
                section_vi_text TEXT,
                status TEXT NOT NULL DEFAULT 'DRAFT',
                created_by INTEGER NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                submitted_at TEXT,
                submitted_to_board_at TEXT,
                board_saved_at TEXT,
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS cash_control_voucher_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                voucher_id INTEGER NOT NULL,
                group_type TEXT NOT NULL,
                item_order INTEGER NOT NULL DEFAULT 0,
                item_name TEXT NOT NULL,
                document_amount REAL NOT NULL DEFAULT 0,
                ksnb_check_type TEXT NOT NULL DEFAULT 'MATCH',
                ksnb_checked_amount REAL,
                ksnb_note TEXT,
                FOREIGN KEY(voucher_id) REFERENCES cash_control_vouchers(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_voucher_signatures (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                voucher_id INTEGER NOT NULL,
                signer_id INTEGER NOT NULL,
                signer_name TEXT NOT NULL,
                signer_role TEXT NOT NULL,
                signature_text TEXT NOT NULL,
                signed_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(voucher_id) REFERENCES cash_control_vouchers(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS cash_control_voucher_routes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                voucher_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                from_user_id INTEGER,
                to_user_id INTEGER,
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(voucher_id) REFERENCES cash_control_vouchers(id) ON DELETE CASCADE
            );      
            """
        )
        for sql in [
            "ALTER TABLE cash_control_vouchers ADD COLUMN current_handler INTEGER",
            "ALTER TABLE cash_control_vouchers ADD COLUMN section_iv_result TEXT",
            "ALTER TABLE cash_control_vouchers ADD COLUMN section_iv_note TEXT",
            "ALTER TABLE cash_control_vouchers ADD COLUMN section_v_text TEXT",
            "ALTER TABLE cash_control_vouchers ADD COLUMN section_vi_text TEXT",
            "ALTER TABLE cash_control_vouchers ADD COLUMN submitted_to_board_at TEXT",
            "ALTER TABLE cash_control_vouchers ADD COLUMN board_saved_at TEXT",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_bank_ml_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_qr_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_pos_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_non_cash_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_report_non_cash_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_employee_summaries ADD COLUMN accounting_non_cash_diff_amount REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_accounting_batches ADD COLUMN accounting_non_cash_diff_total REAL NOT NULL DEFAULT 0",
            "ALTER TABLE cash_control_his_entries ADD COLUMN huy_phieu REAL NOT NULL DEFAULT 0",
        ]:
            try:
                conn.execute(sql)
            except Exception:
                pass

        conn.commit()


def compute_totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    fields = [
        "patient_count",
        "txn_count",
        "tien_thu",
        "tam_ung",
        "tien_chi",
        "mien_giam",
        "huy_phieu",
        "thuc_thu",
        "tien_mat",
        "chuyen_khoan",
        "qr",
        "pos",
    ]
    totals: dict[str, float] = {field: 0.0 for field in fields}
    for row in rows:
        for field in fields:
            totals[field] += money_value(row.get(field))
    totals["khong_bang_tien_mat"] = totals["chuyen_khoan"] + totals["qr"] + totals["pos"]
    return totals


def load_latest_his_context(selected_batch_id: int | None = None) -> dict[str, Any]:
    with get_conn() as conn:
        if selected_batch_id:
            latest_batch = conn.execute(
                """
                SELECT b.*, u.full_name AS creator_name
                FROM cash_control_batches b
                JOIN users u ON u.id = b.created_by
                WHERE b.id = ?
                """,
                (selected_batch_id,),
            ).fetchone()
        else:
            latest_batch = conn.execute(
                """
                SELECT b.*, u.full_name AS creator_name
                FROM cash_control_batches b
                JOIN users u ON u.id = b.created_by
                ORDER BY b.id DESC
                LIMIT 1
                """
            ).fetchone()

        if not latest_batch:
            return {
                "selected_batch": None,
                "summaries": [],
                "totals": {},
            }

        rows = conn.execute(
            """
            SELECT *,
                   COALESCE(chuyen_khoan, 0) + COALESCE(qr, 0) + COALESCE(pos, 0) AS khong_bang_tien_mat
            FROM cash_control_employee_summaries
            WHERE batch_id = ?
            ORDER BY employee_code, employee_name
            """,
            (latest_batch["id"],),
        ).fetchall()

    summary_dicts = [dict(row) for row in rows]
    totals = compute_totals(summary_dicts)
    totals.setdefault("patient_count", 0)
    totals.setdefault("txn_count", 0)
    totals.setdefault("khong_bang_tien_mat", 0)

    return {
        "selected_batch": latest_batch,
        "summaries": summary_dicts,
        "totals": totals,
    }


def load_accounting_context(selected_accounting_batch_id: int | None = None) -> dict[str, Any]:
    ensure_cash_control_tables()

    with get_conn() as conn:
        accounting_batches = conn.execute(
            """
            SELECT b.*, u.full_name AS creator_name
            FROM cash_control_accounting_batches b
            JOIN users u ON u.id = b.created_by
            ORDER BY b.id DESC
            """
        ).fetchall()

        selected_accounting_batch = None
        if selected_accounting_batch_id:
            selected_accounting_batch = conn.execute(
                """
                SELECT *
                FROM cash_control_accounting_batches
                WHERE id = ?
                """,
                (selected_accounting_batch_id,),
            ).fetchone()

        if not selected_accounting_batch:
            selected_accounting_batch = accounting_batches[0] if accounting_batches else None
        accounting_summaries = []
        accounting_employee_summaries = []

        if selected_accounting_batch:
            accounting_summaries = conn.execute(
                """
                SELECT *
                FROM cash_control_accounting_summaries
                WHERE batch_id = ?
                ORDER BY id
                """,
                (selected_accounting_batch["id"],),
            ).fetchall()

            accounting_employee_summaries = conn.execute(
                """
                SELECT *
                FROM cash_control_accounting_employee_summaries
                WHERE batch_id = ?
                ORDER BY employee_code, employee_name
                """,
                (selected_accounting_batch["id"],),
            ).fetchall()

    accounting_dicts = [dict(row) for row in accounting_summaries]
    raw_employee_dicts = [dict(row) for row in accounting_employee_summaries]
    employee_dicts = [
        row
        for row in raw_employee_dicts
        if not normalize_text(row.get("employee_name")).startswith("tong thu")
    ]
    accounting_non_cash_diff_total = money_value(
        selected_accounting_batch["accounting_non_cash_diff_total"]
        if selected_accounting_batch
        else 0
    )
    accounting_employee_totals = {
        "accounting_cash_amount": sum(money_value(row.get("accounting_cash_amount")) for row in employee_dicts),
        "accounting_bank_ml_amount": sum(money_value(row.get("accounting_bank_ml_amount")) for row in employee_dicts),
        "accounting_qr_amount": sum(money_value(row.get("accounting_qr_amount")) for row in employee_dicts),
        "accounting_pos_amount": sum(money_value(row.get("accounting_pos_amount")) for row in employee_dicts),
        "accounting_non_cash_amount": sum(money_value(row.get("accounting_non_cash_amount")) for row in employee_dicts),
        "accounting_report_non_cash_amount": sum(money_value(row.get("accounting_report_non_cash_amount")) for row in employee_dicts),
        "accounting_non_cash_diff_amount": accounting_non_cash_diff_total,
    }

    return {
        "accounting_batches": accounting_batches,
        "selected_accounting_batch": selected_accounting_batch,
        "accounting_summaries": accounting_dicts,
        "accounting_summary_map": {row["item_name"]: row for row in accounting_dicts},
        "accounting_employee_summaries": employee_dicts,
        "accounting_employee_totals": accounting_employee_totals,
    }


def load_cashbook_context(selected_cashbook_batch_id: int | None = None) -> dict[str, Any]:
    ensure_cash_control_tables()

    with get_conn() as conn:
        cashbook_batches = conn.execute(
            """
            SELECT b.*, u.full_name AS creator_name
            FROM cash_control_cashbook_batches b
            JOIN users u ON u.id = b.created_by
            ORDER BY b.id DESC
            """
        ).fetchall()

        selected_cashbook_batch = None
        if selected_cashbook_batch_id:
            selected_cashbook_batch = conn.execute(
                """
                SELECT *
                FROM cash_control_cashbook_batches
                WHERE id = ?
                """,
                (selected_cashbook_batch_id,),
            ).fetchone()

        if not selected_cashbook_batch:
            selected_cashbook_batch = cashbook_batches[0] if cashbook_batches else None
        receipt_rows = []
        expense_rows = []

        if selected_cashbook_batch:
            receipt_rows = conn.execute(
                """
                SELECT *
                FROM cash_control_cashbook_summaries
                WHERE batch_id = ? AND summary_type = 'RECEIPT'
                ORDER BY id
                """,
                (selected_cashbook_batch["id"],),
            ).fetchall()
            expense_rows = conn.execute(
                """
                SELECT *
                FROM cash_control_cashbook_summaries
                WHERE batch_id = ? AND summary_type = 'EXPENSE'
                ORDER BY id
                """,
                (selected_cashbook_batch["id"],),
            ).fetchall()

    receipt_dicts = [dict(row) for row in receipt_rows]
    expense_dicts = [dict(row) for row in expense_rows]

    receipt_count = sum(int(row.get("voucher_count") or 0) for row in receipt_dicts)
    expense_count = sum(int(row.get("voucher_count") or 0) for row in expense_dicts)
    refund_kcb_count = sum(
        int(row.get("voucher_count") or 0)
        for row in expense_dicts
        if row.get("item_name") == "Chi hoàn tiền Khám chữa bệnh"
    )

    return {
        "cashbook_batches": cashbook_batches,
        "selected_cashbook_batch": selected_cashbook_batch,
        "cashbook_receipts": receipt_dicts,
        "cashbook_expenses": expense_dicts,
        "cashbook_receipt_map": {row["item_name"]: row for row in receipt_dicts},
        "cashbook_expense_map": {row["item_name"]: row for row in expense_dicts},
        "cashbook_totals": {
            "receipt_count": receipt_count,
            "expense_count": expense_count,
            "refund_kcb_count": refund_kcb_count,
            "receipt_total": sum(money_value(row.get("amount")) for row in receipt_dicts),
            "expense_total": sum(money_value(row.get("amount")) for row in expense_dicts),
        },
    }


def money_text(value: Any) -> str:
    return f"{money_value(value):,.0f}".replace(",", ".")

def load_cash_control_signature_users() -> dict[str, Any]:
    head_role_codes = list(HEAD_ROLES)
    board_role_codes = list(BOARD_ROLES)

    with get_conn() as conn:
        head_users = []
        board_users = []

        if head_role_codes:
            placeholders = ",".join(["?"] * len(head_role_codes))
            head_users = conn.execute(
                f"""
                SELECT id, full_name, position_title, role_code
                FROM users
                WHERE is_active = 1
                  AND role_code IN ({placeholders})
                ORDER BY role_code, full_name
                """,
                head_role_codes,
            ).fetchall()

        if board_role_codes:
            placeholders = ",".join(["?"] * len(board_role_codes))
            board_users = conn.execute(
                f"""
                SELECT id, full_name, position_title, role_code
                FROM users
                WHERE is_active = 1
                  AND role_code IN ({placeholders})
                ORDER BY role_code, full_name
                """,
                board_role_codes,
            ).fetchall()

    return {
        "cash_control_head_users": head_users,
        "cash_control_board_users": board_users,
    }


def build_control_voucher_context(
    totals: dict[str, Any],
    cashbook_context: dict[str, Any],
    accounting_context: dict[str, Any],
) -> dict[str, Any]:
    cashbook_receipt_map = cashbook_context.get("cashbook_receipt_map", {})
    cashbook_expense_map = cashbook_context.get("cashbook_expense_map", {})
    accounting_summary_map = accounting_context.get("accounting_summary_map", {})

    control_revenue_rows = []
    for item_name in CASH_CONTROL_REVENUE_ITEMS:
        accounting_amount = money_value((accounting_summary_map.get(item_name) or {}).get("amount"))

        if item_name == "Thu tiền khám chữa bệnh":
            ksnb_amount = money_value(totals.get("tien_mat"))
        else:
            ksnb_amount = money_value((cashbook_receipt_map.get(item_name) or {}).get("amount"))

        diff_amount = ksnb_amount - accounting_amount
        if abs(diff_amount) >= 1:
            diff_note = f"Chênh lệch Ban KSNB kiểm tra - Báo cáo Kế toán: {money_text(diff_amount)}."
        else:
            diff_note = "Khớp số liệu giữa Ban KSNB kiểm tra và Báo cáo Kế toán."

        control_revenue_rows.append(
            {
                "item_name": item_name,
                "document_amount": accounting_amount,
                "ksnb_amount": ksnb_amount,
                "diff_amount": diff_amount,
                "note": diff_note,
            }
        )

    control_expense_rows = []
    for item_name in CASH_CONTROL_EXPENSE_ITEMS:
        cashbook_amount = money_value((cashbook_expense_map.get(item_name) or {}).get("amount"))

        control_expense_rows.append(
            {
                "item_name": item_name,
                "document_amount": cashbook_amount,
                "ksnb_amount": cashbook_amount,
                "diff_amount": 0.0,
                "note": "",
            }
        )

    return {
        "control_revenue_rows": control_revenue_rows,
        "control_expense_rows": control_expense_rows,
        "control_revenue_total": {
            "document_amount": sum(row["document_amount"] for row in control_revenue_rows),
            "ksnb_amount": sum(row["ksnb_amount"] for row in control_revenue_rows),
            "diff_amount": sum(row["diff_amount"] for row in control_revenue_rows),
        },
        "control_expense_total": {
            "document_amount": sum(row["document_amount"] for row in control_expense_rows),
            "ksnb_amount": sum(row["ksnb_amount"] for row in control_expense_rows),
            "diff_amount": sum(row["diff_amount"] for row in control_expense_rows),
        },
    }


def cash_control_signature_text(user: dict) -> str:
    return f"Đã ký số nội bộ\n{user['full_name']}\n{user['position_title']}"


def cash_control_no_signature_text(user: dict) -> str:
    return f"Không ký số - in tên trên Phiếu\n{user['full_name']}\n{user['position_title']}"


def cash_control_item_check_amount(item: dict) -> float:
    if item.get("ksnb_check_type") == "REJECTED":
        return 0.0

    if item.get("ksnb_check_type") == "DIFFERENT":
        return money_value(item.get("ksnb_checked_amount"))

    if item.get("group_type") == "REVENUE":
        return money_value(item.get("ksnb_checked_amount"))

    return money_value(item.get("document_amount"))


def fetch_full_cash_control_voucher(voucher_id: int) -> dict[str, Any]:
    ensure_cash_control_tables()

    with get_conn() as conn:
        voucher = conn.execute(
            """
            SELECT v.*,
                   creator.full_name AS creator_name,
                   creator.position_title AS creator_position,
                   head.full_name AS head_name,
                   head.position_title AS head_position,
                   board.full_name AS board_name,
                   board.position_title AS board_position
            FROM cash_control_vouchers v
            JOIN users creator ON creator.id = v.created_by
            LEFT JOIN users head ON head.id = v.head_user_id
            LEFT JOIN users board ON board.id = v.board_user_id
            WHERE v.id = ?
            """,
            (voucher_id,),
        ).fetchone()

        if not voucher:
            return {}

        items = conn.execute(
            """
            SELECT *
            FROM cash_control_voucher_items
            WHERE voucher_id = ?
            ORDER BY
                CASE group_type
                    WHEN 'REVENUE' THEN 1
                    WHEN 'EXPENSE' THEN 2
                    ELSE 3
                END,
                item_order,
                id
            """,
            (voucher_id,),
        ).fetchall()

        signatures = conn.execute(
            """
            SELECT *
            FROM cash_control_voucher_signatures
            WHERE voucher_id = ?
            ORDER BY id
            """,
            (voucher_id,),
        ).fetchall()

        routes = conn.execute(
            """
            SELECT r.*,
                   fu.full_name AS from_name,
                   tu.full_name AS to_name
            FROM cash_control_voucher_routes r
            LEFT JOIN users fu ON fu.id = r.from_user_id
            LEFT JOIN users tu ON tu.id = r.to_user_id
            WHERE r.voucher_id = ?
            ORDER BY r.id
            """,
            (voucher_id,),
        ).fetchall()

    item_dicts = [
        dict(row)
        for row in items
        if str(row["item_name"] or "").strip() != "Thu tiền Quầy kính"
    ]
    revenue_items = [row for row in item_dicts if row["group_type"] == "REVENUE"]
    expense_items = [row for row in item_dicts if row["group_type"] == "EXPENSE"]

    document_total = sum(money_value(row["document_amount"]) for row in item_dicts)
    ksnb_total = sum(cash_control_item_check_amount(row) for row in item_dicts)

    return {
        "voucher": voucher,
        "items": item_dicts,
        "revenue_items": revenue_items,
        "expense_items": expense_items,
        "signatures": signatures,
        "routes": routes,
        "document_total": document_total,
        "ksnb_total": ksnb_total,
        "diff_total": document_total - ksnb_total,
    }


def can_view_cash_control_voucher(user: dict, voucher: Any) -> bool:
    if not voucher:
        return False

    if user.get("role_code") == "ADMIN":
        return True

    if voucher["created_by"] == user.get("id"):
        return True

    if user.get("role_code") in HEAD_ROLES:
        return True

    if user.get("role_code") in BOARD_ROLES and voucher["status"] in {
        "SUBMITTED_TO_BOARD",
        "BOARD_VIEWED",
        "BOARD_SAVED",
    }:
        return True

    return False

def save_cash_control_voucher_record(
    request: Request,
    *,
    route_mode: str,
    head_user_id: int | None,
    voucher_date: str,
    his_batch_id: int | None,
    cashbook_batch_id: int | None,
    accounting_batch_id: int | None,
) -> int:
    ensure_cash_control_tables()

    user = request.state.user
    his_context = load_latest_his_context(his_batch_id)
    cashbook_context = load_cashbook_context(cashbook_batch_id)
    accounting_context = load_accounting_context(accounting_batch_id)
    control_context = build_control_voucher_context(
        his_context["totals"],
        cashbook_context,
        accounting_context,
    )

    selected_his = his_context.get("selected_batch")
    selected_cashbook = cashbook_context.get("selected_cashbook_batch")
    selected_accounting = accounting_context.get("selected_accounting_batch")
    cashbook_totals = cashbook_context.get("cashbook_totals", {})
    clean_voucher_date = format_vn_control_date(voucher_date)

    title = "Phiếu kiểm soát thu, chi tiền mặt"
    if clean_voucher_date:
        title = f"{title} ngày {clean_voucher_date}"

    section_iv_result = "Hồ sơ đủ điều kiện trình phê duyệt."
    section_iv_note = ""
    if (
        abs(control_context["control_revenue_total"]["diff_amount"]) >= 1
        or abs(control_context["control_expense_total"]["diff_amount"]) >= 1
    ):
        section_iv_result = "Hồ sơ cần bổ sung, chỉnh sửa trước khi trình phê duyệt."
        section_iv_note = (
            "Phiếu có phát sinh chênh lệch giữa Báo cáo Kế toán và số liệu Ban KSNB kiểm tra; "
            "đề nghị xem chi tiết tại Mục III."
        )

    source_note = []
    if selected_his:
        source_note.append(f"HIS: #{selected_his['id']} - {selected_his['original_filename']}")
    if selected_cashbook:
        source_note.append(f"Sổ quỹ: #{selected_cashbook['id']} - {selected_cashbook['original_filename']}")
    if selected_accounting:
        source_note.append(f"Kế toán: #{selected_accounting['id']} - {selected_accounting['original_filename']}")

    section_v_text = "Nguồn dữ liệu kiểm soát: " + "; ".join(source_note) if source_note else ""

    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO cash_control_vouchers(
                title, voucher_date, sign_mode, route_mode, head_user_id, board_user_id, current_handler,
                his_batch_id, cashbook_batch_id, accounting_batch_id,
                receipt_count, expense_count, refund_kcb_count,
                revenue_document_total, revenue_ksnb_total, revenue_diff_total,
                expense_document_total, expense_ksnb_total, expense_diff_total,
                section_iv_result, section_iv_note, section_v_text, section_vi_text,
                status, created_by
            )
            VALUES (?, ?, 'INTERNAL_DIGITAL', ?, ?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', 'DRAFT', ?)
            """,
            (
                title,
                clean_voucher_date,
                route_mode,
                head_user_id,
                head_user_id if route_mode == "THROUGH_HEAD" else None,
                his_batch_id,
                cashbook_batch_id,
                accounting_batch_id,
                int(cashbook_totals.get("receipt_count", 0) or 0),
                int(cashbook_totals.get("expense_count", 0) or 0),
                int(cashbook_totals.get("refund_kcb_count", 0) or 0),
                money_value(control_context["control_revenue_total"]["document_amount"]),
                money_value(control_context["control_revenue_total"]["ksnb_amount"]),
                money_value(control_context["control_revenue_total"]["diff_amount"]),
                money_value(control_context["control_expense_total"]["document_amount"]),
                money_value(control_context["control_expense_total"]["ksnb_amount"]),
                money_value(control_context["control_expense_total"]["diff_amount"]),
                section_iv_result,
                section_iv_note,
                section_v_text,
                user["id"],
            ),
        )
        voucher_id = cur.lastrowid
        code = f"TM-KSNB-{voucher_id:05d}"
        conn.execute("UPDATE cash_control_vouchers SET code = ? WHERE id = ?", (code, voucher_id))

        item_order = 1
        for row in control_context["control_revenue_rows"]:
            check_type = "MATCH"
            checked_amount = money_value(row["ksnb_amount"])
            if abs(row["diff_amount"]) >= 1:
                check_type = "DIFFERENT"

            conn.execute(
                """
                INSERT INTO cash_control_voucher_items(
                    voucher_id, group_type, item_order, item_name, document_amount,
                    ksnb_check_type, ksnb_checked_amount, ksnb_note
                )
                VALUES (?, 'REVENUE', ?, ?, ?, ?, ?, ?)
                """,
                (
                    voucher_id,
                    item_order,
                    row["item_name"],
                    money_value(row["document_amount"]),
                    check_type,
                    checked_amount,
                    row["note"],
                ),
            )
            item_order += 1

        item_order = 1
        for row in control_context["control_expense_rows"]:
            check_type = "MATCH"
            checked_amount = None

            conn.execute(
                """
                INSERT INTO cash_control_voucher_items(
                    voucher_id, group_type, item_order, item_name, document_amount,
                    ksnb_check_type, ksnb_checked_amount, ksnb_note
                )
                VALUES (?, 'EXPENSE', ?, ?, ?, ?, ?, ?)
                """,
                (
                    voucher_id,
                    item_order,
                    row["item_name"],
                    money_value(row["document_amount"]),
                    check_type,
                    checked_amount,
                    row["note"],
                ),
            )
            item_order += 1

        conn.execute(
            """
            INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
            VALUES (?, 'TAO_PHIEU', ?, ?, ?)
            """,
            (
                voucher_id,
                user["id"],
                head_user_id if route_mode == "THROUGH_HEAD" else None,
                "Tạo Phiếu kiểm soát thu, chi tiền mặt",
            ),
        )
        conn.commit()

    return int(voucher_id)

def update_cash_control_voucher_basic_record(
    request: Request,
    *,
    voucher_id: int,
    route_mode: str,
    head_user_id: int | None,
    voucher_date: str,
    his_batch_id: int | None,
    cashbook_batch_id: int | None,
    accounting_batch_id: int | None,
) -> None:
    ensure_cash_control_tables()

    user = request.state.user
    his_context = load_latest_his_context(his_batch_id)
    cashbook_context = load_cashbook_context(cashbook_batch_id)
    accounting_context = load_accounting_context(accounting_batch_id)
    control_context = build_control_voucher_context(
        his_context["totals"],
        cashbook_context,
        accounting_context,
    )

    selected_his = his_context.get("selected_batch")
    selected_cashbook = cashbook_context.get("selected_cashbook_batch")
    selected_accounting = accounting_context.get("selected_accounting_batch")
    cashbook_totals = cashbook_context.get("cashbook_totals", {})
    clean_voucher_date = format_vn_control_date(voucher_date)

    title = "Phiếu kiểm soát thu, chi tiền mặt"
    if clean_voucher_date:
        title = f"{title} ngày {clean_voucher_date}"

    section_iv_result = "Hồ sơ đủ điều kiện trình phê duyệt."
    section_iv_note = ""
    if (
        abs(control_context["control_revenue_total"]["diff_amount"]) >= 1
        or abs(control_context["control_expense_total"]["diff_amount"]) >= 1
    ):
        section_iv_result = "Hồ sơ cần bổ sung, chỉnh sửa trước khi trình phê duyệt."
        section_iv_note = (
            "Phiếu có phát sinh chênh lệch giữa Báo cáo Kế toán và số liệu Ban KSNB kiểm tra; "
            "đề nghị xem chi tiết tại Mục III."
        )

    source_note = []
    if selected_his:
        source_note.append(f"HIS: #{selected_his['id']} - {selected_his['original_filename']}")
    if selected_cashbook:
        source_note.append(f"Sổ quỹ: #{selected_cashbook['id']} - {selected_cashbook['original_filename']}")
    if selected_accounting:
        source_note.append(f"Kế toán: #{selected_accounting['id']} - {selected_accounting['original_filename']}")

    section_v_text = "Nguồn dữ liệu kiểm soát: " + "; ".join(source_note) if source_note else ""

    with get_conn() as conn:
        existing_voucher = conn.execute(
            """
            SELECT section_iv_result, section_iv_note, section_v_text, section_vi_text
            FROM cash_control_vouchers
            WHERE id = ?
            """,
            (voucher_id,),
        ).fetchone()

        if existing_voucher:
            existing_section_iv_result = str(existing_voucher["section_iv_result"] or "").strip()
            existing_section_iv_note = str(existing_voucher["section_iv_note"] or "").strip()
            existing_section_v_text = str(existing_voucher["section_v_text"] or "").strip()

            if existing_section_iv_result:
                section_iv_result = existing_section_iv_result
            if existing_section_iv_note:
                section_iv_note = existing_section_iv_note
            if existing_section_v_text and not existing_section_v_text.startswith("Nguồn dữ liệu kiểm soát:"):
                section_v_text = existing_section_v_text

        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET title = ?,
                voucher_date = ?,
                route_mode = ?,
                head_user_id = ?,
                current_handler = ?,
                his_batch_id = ?,
                cashbook_batch_id = ?,
                accounting_batch_id = ?,
                receipt_count = ?,
                expense_count = ?,
                refund_kcb_count = ?,
                revenue_document_total = ?,
                revenue_ksnb_total = ?,
                revenue_diff_total = ?,
                expense_document_total = ?,
                expense_ksnb_total = ?,
                expense_diff_total = ?,
                section_iv_result = ?,
                section_iv_note = ?,
                section_v_text = ?
            WHERE id = ?
            """,
            (
                title,
                clean_voucher_date,
                route_mode,
                head_user_id,
                head_user_id if route_mode == "THROUGH_HEAD" else None,
                his_batch_id,
                cashbook_batch_id,
                accounting_batch_id,
                int(cashbook_totals.get("receipt_count", 0) or 0),
                int(cashbook_totals.get("expense_count", 0) or 0),
                int(cashbook_totals.get("refund_kcb_count", 0) or 0),
                money_value(control_context["control_revenue_total"]["document_amount"]),
                money_value(control_context["control_revenue_total"]["ksnb_amount"]),
                money_value(control_context["control_revenue_total"]["diff_amount"]),
                money_value(control_context["control_expense_total"]["document_amount"]),
                money_value(control_context["control_expense_total"]["ksnb_amount"]),
                money_value(control_context["control_expense_total"]["diff_amount"]),
                section_iv_result,
                section_iv_note,
                section_v_text,
                voucher_id,
            ),
        )

        conn.execute(
            "DELETE FROM cash_control_voucher_items WHERE voucher_id = ?",
            (voucher_id,),
        )

        item_order = 1
        for row in control_context["control_revenue_rows"]:
            check_type = "MATCH"
            checked_amount = money_value(row["ksnb_amount"])
            if abs(row["diff_amount"]) >= 1:
                check_type = "DIFFERENT"

            conn.execute(
                """
                INSERT INTO cash_control_voucher_items(
                    voucher_id, group_type, item_order, item_name, document_amount,
                    ksnb_check_type, ksnb_checked_amount, ksnb_note
                )
                VALUES (?, 'REVENUE', ?, ?, ?, ?, ?, ?)
                """,
                (
                    voucher_id,
                    item_order,
                    row["item_name"],
                    money_value(row["document_amount"]),
                    check_type,
                    checked_amount,
                    row["note"],
                ),
            )
            item_order += 1

        item_order = 1
        for row in control_context["control_expense_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_voucher_items(
                    voucher_id, group_type, item_order, item_name, document_amount,
                    ksnb_check_type, ksnb_checked_amount, ksnb_note
                )
                VALUES (?, 'EXPENSE', ?, ?, ?, 'MATCH', NULL, ?)
                """,
                (
                    voucher_id,
                    item_order,
                    row["item_name"],
                    money_value(row["document_amount"]),
                    row["note"],
                ),
            )
            item_order += 1

        conn.execute(
            """
            INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
            VALUES (?, 'CAP_NHAT_NHAP', ?, ?, ?)
            """,
            (
                voucher_id,
                user["id"],
                head_user_id if route_mode == "THROUGH_HEAD" else None,
                "Cập nhật thông tin ban đầu của Phiếu nháp",
            ),
        )
        conn.commit()

def build_reconcile_context(
    totals: dict[str, Any],
    summaries: list[dict[str, Any]],
    cashbook_context: dict[str, Any],
    accounting_context: dict[str, Any],
    selected_his_batch_id: int | None = None,
) -> dict[str, Any]:
    cashbook_receipt_map = cashbook_context.get("cashbook_receipt_map", {})
    cashbook_expense_map = cashbook_context.get("cashbook_expense_map", {})
    accounting_summary_map = accounting_context.get("accounting_summary_map", {})
    accounting_employee_totals = accounting_context.get("accounting_employee_totals", {})
    accounting_non_cash_diff_total = money_value(
        accounting_employee_totals.get("accounting_non_cash_diff_amount")
    )

    revenue_matrix = []
    kcb_not_in_his_rows = []
    
    for item_name in RECONCILE_REVENUE_ITEMS:
        his_amount = 0.0
        cashbook_amount = 0.0
        accounting_amount = money_value((accounting_summary_map.get(item_name) or {}).get("amount"))

        if item_name == "Thu tiền khám chữa bệnh":
            his_cash_amount = money_value(totals.get("tien_mat"))
            optical_cash_amount = money_value(
                (cashbook_receipt_map.get("Thu tiền Quầy kính") or {}).get("amount")
            )
            cashbook_kcb_amount = money_value(
                (cashbook_receipt_map.get("Thu tiền khám chữa bệnh") or {}).get("amount")
            )
            his_amount = his_cash_amount
            cashbook_amount = cashbook_kcb_amount + optical_cash_amount
        else:
            cashbook_amount = money_value((cashbook_receipt_map.get(item_name) or {}).get("amount"))

        diff_his_cashbook = his_amount - cashbook_amount if item_name == "Thu tiền khám chữa bệnh" else 0.0
        diff_ksnb_cashbook_minus_accounting_report_diff = (
            diff_his_cashbook - accounting_non_cash_diff_total
            if item_name == "Thu tiền khám chữa bệnh"
            else 0.0
        )
        diff_his_accounting = his_amount - accounting_amount if item_name == "Thu tiền khám chữa bệnh" else 0.0
        diff_cashbook_accounting = cashbook_amount - accounting_amount

        revenue_matrix.append(
            {
                "item_name": item_name,
                "his_amount": his_amount,
                "cashbook_amount": cashbook_amount,
                "accounting_amount": accounting_amount,
                "diff_his_cashbook": diff_his_cashbook,
                "diff_ksnb_cashbook_minus_accounting_report_diff": diff_ksnb_cashbook_minus_accounting_report_diff,
                "diff_his_accounting": diff_his_accounting,
                "diff_cashbook_accounting": diff_cashbook_accounting,
            }
        )

    his_employee_name_set = {
        normalize_text(row.get("employee_name"))
        for row in summaries
        if normalize_text(row.get("employee_name"))
    }

    cashbook_kcb_details = []
    selected_cashbook_batch = cashbook_context.get("selected_cashbook_batch")
    selected_cashbook_batch_id = selected_cashbook_batch["id"] if selected_cashbook_batch else None
    allowed_cashbook_kcb_codes = {"KHACHKCB", "KLKSK", "QUAYKINH", "THAMMY"}

    if selected_his_batch_id and selected_cashbook_batch_id:
        with get_conn() as conn:
            cashbook_kcb_details = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT voucher_no, voucher_date, description, amount, credit_code
                    FROM cash_control_cashbook_entries
                    WHERE batch_id = ?
                      AND doc_type = 'PT'
                      AND UPPER(COALESCE(credit_code, '')) IN ('KHACHKCB', 'KLKSK', 'QUAYKINH', 'THAMMY')
                    ORDER BY voucher_date, voucher_no, id
                    """,
                    (selected_cashbook_batch_id,),
                ).fetchall()
            ]

    used_cashbook_detail_indexes = set()

    if selected_his_batch_id and selected_cashbook_batch_id:
        for accounting_row in accounting_context.get("accounting_employee_summaries", []):
            accounting_name = str(accounting_row.get("employee_name") or "").strip()
            accounting_cash_amount = money_value(accounting_row.get("accounting_cash_amount"))
            accounting_note = str(accounting_row.get("note") or "").strip()

            if not accounting_name or accounting_cash_amount <= 0:
                continue

            if normalize_text(accounting_name) in his_employee_name_set:
                continue

            matched_cashbook_detail = None
            for detail_index, detail in enumerate(cashbook_kcb_details):
                if detail_index in used_cashbook_detail_indexes:
                    continue

                cashbook_code = str(detail.get("credit_code") or "").strip().upper()
                if cashbook_code not in allowed_cashbook_kcb_codes:
                    continue

                if abs(money_value(detail.get("amount")) - accounting_cash_amount) < 1:
                    matched_cashbook_detail = detail
                    used_cashbook_detail_indexes.add(detail_index)
                    break

            if not matched_cashbook_detail:
                continue

            cashbook_amount = money_value(matched_cashbook_detail.get("amount"))
            voucher_no = str(matched_cashbook_detail.get("voucher_no") or "").strip()
            voucher_date = str(matched_cashbook_detail.get("voucher_date") or "").strip()
            cashbook_description = str(matched_cashbook_detail.get("description") or "").strip()
            cashbook_code = str(matched_cashbook_detail.get("credit_code") or "").strip().upper()

            content_parts = [
                "Thu tiền mặt KCB có trên Sổ quỹ/Kế toán nhưng không có trong tổng hợp HIS",
                f"Tên ghi nhận trên Bảng kê Kế toán: {accounting_name}",
                f"Mã Sổ quỹ: {cashbook_code}",
            ]

            if voucher_no:
                content_parts.append(f"Số phiếu: {voucher_no}")
            if voucher_date:
                content_parts.append(f"Ngày: {voucher_date}")
            if cashbook_description:
                content_parts.append(f"Diễn giải Sổ quỹ: {cashbook_description}")
            if accounting_note:
                content_parts.append(f"Ghi chú Kế toán: {accounting_note}")

            note_text = (
                "Khoản thu tiền mặt KCB có trong Báo cáo Kế toán và có trong Sổ quỹ tiền mặt "
                "với mã thuộc nhóm KHACHKCB/KLKSK/QUAYKINH, nhưng tên ghi nhận không có trong "
                "tổng hợp doanh thu tiền mặt trên HIS."
            )

            kcb_not_in_his_rows.append(
                {
                    "content": "; ".join(content_parts),
                    "cashbook_amount": cashbook_amount,
                    "accounting_amount": accounting_cash_amount,
                    "his_amount": 0.0,
                    "diff_amount": cashbook_amount,
                    "note": note_text,
                }
            )

    expense_matrix = []
    for item_name in CASH_CONTROL_EXPENSE_ITEMS:
        cashbook_amount = money_value((cashbook_expense_map.get(item_name) or {}).get("amount"))
        accounting_amount = money_value((accounting_summary_map.get(item_name) or {}).get("amount"))
        
        expense_matrix.append(
            {
                "item_name": item_name,
                "cashbook_amount": cashbook_amount,
                "accounting_amount": accounting_amount,
                "diff_amount": cashbook_amount - accounting_amount,
            }
        )

    employee_accounting_map = {
        row["employee_code"]: row for row in accounting_context.get("accounting_employee_summaries", [])
    }

    employee_reconcile_rows = []
    employee_non_cash_reconcile_rows = []
    for row in summaries:
        employee_code = str(row.get("employee_code") or "").strip()
        accounting_row = employee_accounting_map.get(employee_code)
        his_cash_amount = money_value(row.get("tien_mat"))
        accounting_cash_amount = money_value((accounting_row or {}).get("accounting_cash_amount"))
        diff_amount = his_cash_amount - accounting_cash_amount

        if abs(diff_amount) >= 1:
            employee_reconcile_rows.append(
                {
                    "employee_code": employee_code,
                    "employee_name": row.get("employee_name"),
                    "his_cash_amount": his_cash_amount,
                    "accounting_cash_amount": accounting_cash_amount,
                    "diff_amount": diff_amount,
                }
            )

        his_non_cash_amount = (
            money_value(row.get("chuyen_khoan"))
            + money_value(row.get("qr"))
            + money_value(row.get("pos"))
        )
        accounting_non_cash_amount = money_value((accounting_row or {}).get("accounting_non_cash_amount"))
        accounting_report_non_cash_amount = money_value(
            (accounting_row or {}).get("accounting_report_non_cash_amount")
        )
        diff_his_accounting_non_cash_amount = his_non_cash_amount - accounting_non_cash_amount
        diff_his_accounting_report_non_cash_amount = (
            his_non_cash_amount
            - accounting_report_non_cash_amount
        )

        if (
            abs(diff_his_accounting_non_cash_amount) >= 1
            or abs(diff_his_accounting_report_non_cash_amount) >= 1
        ):
            employee_non_cash_reconcile_rows.append(
                {
                    "employee_code": employee_code,
                    "employee_name": row.get("employee_name"),
                    "his_non_cash_amount": his_non_cash_amount,
                    "accounting_non_cash_amount": accounting_non_cash_amount,
                    "accounting_report_non_cash_amount": accounting_report_non_cash_amount,
                    "diff_his_accounting_non_cash_amount": diff_his_accounting_non_cash_amount,
                    "diff_his_accounting_report_non_cash_amount": diff_his_accounting_report_non_cash_amount,
                }
            )

    receipt_difference_details: list[dict[str, Any]] = []
    discrepant_employee_codes = [
        str(row.get("employee_code") or "").strip()
        for row in employee_reconcile_rows
        if str(row.get("employee_code") or "").strip()
    ]

    if selected_his_batch_id and discrepant_employee_codes:
        placeholders = ",".join("?" for _ in discrepant_employee_codes)
        with get_conn() as conn:
            detail_rows = conn.execute(
                f"""
                SELECT *
                FROM cash_control_his_entries
                WHERE batch_id = ?
                  AND employee_code IN ({placeholders})
                ORDER BY employee_code, patient_code, id
                """,
                [selected_his_batch_id, *discrepant_employee_codes],
            ).fetchall()
        receipt_difference_details = [dict(row) for row in detail_rows]

    non_cash_difference_details: list[dict[str, Any]] = []
    discrepant_non_cash_employee_codes = [
        str(row.get("employee_code") or "").strip()
        for row in employee_non_cash_reconcile_rows
        if str(row.get("employee_code") or "").strip()
    ]

    if selected_his_batch_id and discrepant_non_cash_employee_codes:
        placeholders = ",".join("?" for _ in discrepant_non_cash_employee_codes)
        with get_conn() as conn:
            detail_rows = conn.execute(
                f"""
                SELECT *,
                       COALESCE(chuyen_khoan, 0) + COALESCE(qr, 0) + COALESCE(pos, 0) AS khong_bang_tien_mat
                FROM cash_control_his_entries
                WHERE batch_id = ?
                  AND employee_code IN ({placeholders})
                  AND (COALESCE(chuyen_khoan, 0) + COALESCE(qr, 0) + COALESCE(pos, 0)) > 0
                ORDER BY employee_code, patient_code, id
                """,
                [selected_his_batch_id, *discrepant_non_cash_employee_codes],
            ).fetchall()
        non_cash_difference_details = [dict(row) for row in detail_rows]

    return {
        "reconcile_revenue_matrix": revenue_matrix,
        "kcb_not_in_his_rows": kcb_not_in_his_rows,
        "reconcile_expense_matrix": expense_matrix,
        "employee_reconcile_rows": employee_reconcile_rows,
        "employee_non_cash_reconcile_rows": employee_non_cash_reconcile_rows,
        "receipt_difference_employees": employee_reconcile_rows,
        "receipt_difference_details": receipt_difference_details,
        "non_cash_difference_employees": employee_non_cash_reconcile_rows,
        "non_cash_difference_details": non_cash_difference_details,
    }


@router.get("")
def cash_control_index(request: Request):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    active_main_tab = request.query_params.get("main_tab") or "summary"
    active_summary_tab = request.query_params.get("summary_tab") or "his"

    if user.get("role_code") in BOARD_ROLES:
        active_main_tab = "voucher"
        active_summary_tab = "his"

    selected_his_import_year = query_text(request, "his_import_year")
    selected_his_import_month = query_text(request, "his_import_month")
    selected_his_import_day = query_text(request, "his_import_day")
    selected_cashbook_import_year = query_text(request, "cashbook_import_year")
    selected_cashbook_import_month = query_text(request, "cashbook_import_month")
    selected_cashbook_import_day = query_text(request, "cashbook_import_day")
    selected_accounting_import_year = query_text(request, "accounting_import_year")
    selected_accounting_import_month = query_text(request, "accounting_import_month")
    selected_accounting_import_day = query_text(request, "accounting_import_day")
    selected_cash_voucher_year = query_text(request, "cash_voucher_year")
    selected_cash_voucher_month = query_text(request, "cash_voucher_month")
    selected_cash_voucher_day = query_text(request, "cash_voucher_day")
    editing_voucher_id = query_int(request, "editing_voucher_id")
    if user.get("role_code") in BOARD_ROLES:
        editing_voucher_id = None

    with get_conn() as conn:
        batches = conn.execute(
            """
            SELECT b.*, u.full_name AS creator_name
            FROM cash_control_batches b
            JOIN users u ON u.id = b.created_by
            ORDER BY b.id DESC
            """
        ).fetchall()

        cash_voucher_where = ["1=1"]
        cash_voucher_params: list[Any] = []

        if user.get("role_code") == "ADMIN":
            pass
        elif user.get("role_code") in HEAD_ROLES:
            pass
        elif user.get("role_code") in BOARD_ROLES:
            cash_voucher_where.append("v.status IN ('SUBMITTED_TO_BOARD', 'BOARD_VIEWED', 'BOARD_SAVED')")
        else:
            cash_voucher_where.append("v.created_by = ?")
            cash_voucher_params.append(user["id"])

        if selected_cash_voucher_year:
            cash_voucher_where.append("strftime('%Y', v.created_at) = ?")
            cash_voucher_params.append(selected_cash_voucher_year)

        if selected_cash_voucher_month:
            cash_voucher_where.append("strftime('%m', v.created_at) = ?")
            cash_voucher_params.append(selected_cash_voucher_month.zfill(2))

        if selected_cash_voucher_day:
            cash_voucher_where.append("strftime('%d', v.created_at) = ?")
            cash_voucher_params.append(selected_cash_voucher_day.zfill(2))

        cash_voucher_where_sql = " AND ".join(cash_voucher_where)

        cash_control_vouchers = conn.execute(
            f"""
            SELECT v.*, u.full_name AS creator_name
            FROM cash_control_vouchers v
            JOIN users u ON u.id = v.created_by
            WHERE {cash_voucher_where_sql}
            ORDER BY v.id DESC
            """,
            cash_voucher_params,
        ).fetchall()

        cash_control_voucher_years = conn.execute(
            """
            SELECT DISTINCT strftime('%Y', created_at) AS year
            FROM cash_control_vouchers
            WHERE created_at IS NOT NULL
            ORDER BY year DESC
            """
        ).fetchall()

        editing_cash_voucher = None
        if editing_voucher_id:
            editing_cash_voucher = conn.execute(
                """
                SELECT v.*, u.full_name AS creator_name
                FROM cash_control_vouchers v
                JOIN users u ON u.id = v.created_by
                WHERE v.id = ?
                """,
                (editing_voucher_id,),
            ).fetchone()

            if not can_view_cash_control_voucher(user, editing_cash_voucher):
                editing_cash_voucher = None

    cash_control_voucher_rows = []
    for row in cash_control_vouchers:
        item = dict(row)
        item["is_new_for_current_user"] = False

        if user["role_code"] == "ADMIN" and item["status"] in ("SUBMITTED_TO_HEAD", "SUBMITTED_TO_BOARD"):
            item["is_new_for_current_user"] = True
        elif user["role_code"] in HEAD_ROLES and item["status"] == "SUBMITTED_TO_HEAD" and item["current_handler"] == user["id"]:
            item["is_new_for_current_user"] = True
        elif user["role_code"] in BOARD_ROLES and item["status"] == "SUBMITTED_TO_BOARD" and (
            item["current_handler"] == user["id"]
            or item["board_user_id"] == user["id"]
            or item["current_handler"] is None
        ):
            item["is_new_for_current_user"] = True

        cash_control_voucher_rows.append(item)

    cash_control_new_count = sum(1 for item in cash_control_voucher_rows if item["is_new_for_current_user"])

    cash_control_voucher_months = [f"{i:02d}" for i in range(1, 13)]
    cash_control_voucher_days = [f"{i:02d}" for i in range(1, 32)]

    his_import_filter_choices = import_filter_choices(list(batches))
    filtered_batches = filter_import_batches(
        list(batches),
        selected_year=selected_his_import_year,
        selected_month=selected_his_import_month,
        selected_day=selected_his_import_day,
    )

    selected_his_batch_id = query_int(request, "his_batch_id")
    selected_cashbook_batch_id = query_int(request, "cashbook_batch_id")
    selected_accounting_batch_id = query_int(request, "accounting_batch_id")
    
    if active_main_tab == "voucher" and editing_cash_voucher:
        if selected_his_batch_id is None:
            selected_his_batch_id = editing_cash_voucher["his_batch_id"]
        if selected_cashbook_batch_id is None:
            selected_cashbook_batch_id = editing_cash_voucher["cashbook_batch_id"]
        if selected_accounting_batch_id is None:
            selected_accounting_batch_id = editing_cash_voucher["accounting_batch_id"]    

    his_context = load_latest_his_context(selected_his_batch_id) if active_main_tab in {"reconcile", "voucher"} else {
        "selected_batch": None,
        "summaries": [],
        "totals": {},
    }
    cashbook_context = load_cashbook_context(selected_cashbook_batch_id)
    accounting_context = load_accounting_context(selected_accounting_batch_id)
    reconcile_context = build_reconcile_context(
        his_context["totals"],
        his_context["summaries"],
        cashbook_context,
        accounting_context,
        his_context["selected_batch"]["id"] if his_context["selected_batch"] else None,
    )
    control_voucher_context = build_control_voucher_context(
        his_context["totals"],
        cashbook_context,
        accounting_context,
    )
    signature_user_context = load_cash_control_signature_users()
    cashbook_import_filter_choices = import_filter_choices(list(cashbook_context.get("cashbook_batches", [])))
    accounting_import_filter_choices = import_filter_choices(list(accounting_context.get("accounting_batches", [])))
    filtered_cashbook_batches = filter_import_batches(
        list(cashbook_context.get("cashbook_batches", [])),
        selected_year=selected_cashbook_import_year,
        selected_month=selected_cashbook_import_month,
        selected_day=selected_cashbook_import_day,
    )
    filtered_accounting_batches = filter_import_batches(
        list(accounting_context.get("accounting_batches", [])),
        selected_year=selected_accounting_import_year,
        selected_month=selected_accounting_import_month,
        selected_day=selected_accounting_import_day,
    )

    return templates.TemplateResponse(
        "cash_control_index.html",
        {
            "request": request,
            "user": user,
            "batches": filtered_batches,
            "his_import_years": his_import_filter_choices["years"],
            "his_import_months": his_import_filter_choices["months"],
            "his_import_days": his_import_filter_choices["days"],
            "selected_his_import_year": selected_his_import_year,
            "selected_his_import_month": selected_his_import_month.zfill(2) if selected_his_import_month else "",
            "selected_his_import_day": selected_his_import_day.zfill(2) if selected_his_import_day else "",
            "cashbook_batches_filtered": filtered_cashbook_batches,
            "cashbook_import_years": cashbook_import_filter_choices["years"],
            "cashbook_import_months": cashbook_import_filter_choices["months"],
            "cashbook_import_days": cashbook_import_filter_choices["days"],
            "selected_cashbook_import_year": selected_cashbook_import_year,
            "selected_cashbook_import_month": selected_cashbook_import_month.zfill(2) if selected_cashbook_import_month else "",
            "selected_cashbook_import_day": selected_cashbook_import_day.zfill(2) if selected_cashbook_import_day else "",
            "accounting_batches_filtered": filtered_accounting_batches,
            "accounting_import_years": accounting_import_filter_choices["years"],
            "accounting_import_months": accounting_import_filter_choices["months"],
            "accounting_import_days": accounting_import_filter_choices["days"],
            "selected_accounting_import_year": selected_accounting_import_year,
            "selected_accounting_import_month": selected_accounting_import_month.zfill(2) if selected_accounting_import_month else "",
            "selected_accounting_import_day": selected_accounting_import_day.zfill(2) if selected_accounting_import_day else "",
            "cash_control_status_labels": CASH_CONTROL_STATUS_LABELS,
            "cash_control_vouchers": cash_control_voucher_rows,
            "cash_control_voucher_years": cash_control_voucher_years,
            "cash_control_voucher_months": cash_control_voucher_months,
            "cash_control_new_count": cash_control_new_count,
            "cash_control_voucher_days": cash_control_voucher_days,
            "selected_cash_voucher_year": selected_cash_voucher_year,
            "selected_cash_voucher_month": selected_cash_voucher_month.zfill(2) if selected_cash_voucher_month else "",
            "selected_cash_voucher_day": selected_cash_voucher_day.zfill(2) if selected_cash_voucher_day else "",
            "editing_cash_voucher": editing_cash_voucher,
            "selected_batch": his_context["selected_batch"],
            "summaries": his_context["summaries"],
            "totals": his_context["totals"],
            "error": None,
            "active_main_tab": active_main_tab,
            "active_summary_tab": active_summary_tab,
            "cash_control_revenue_items": CASH_CONTROL_REVENUE_ITEMS,
            "cash_control_expense_items": CASH_CONTROL_EXPENSE_ITEMS,
            "format_vn_sqlite_datetime": format_vn_sqlite_datetime,
            "format_vn_control_date": format_vn_control_date,
            "selected_his_batch_id": selected_his_batch_id or (his_context["selected_batch"]["id"] if his_context["selected_batch"] else None),
            "selected_cashbook_batch_id": selected_cashbook_batch_id or (cashbook_context["selected_cashbook_batch"]["id"] if cashbook_context["selected_cashbook_batch"] else None),
            "selected_accounting_batch_id": selected_accounting_batch_id or (accounting_context["selected_accounting_batch"]["id"] if accounting_context["selected_accounting_batch"] else None),
            **cashbook_context,
            **accounting_context,
            **reconcile_context,
            **control_voucher_context,
            **signature_user_context,
        },
    )

@router.post("/voucher/save")
async def cash_control_voucher_save(
    request: Request,
    voucher_date: str = Form(""),
    sign_mode: str = Form("INTERNAL_DIGITAL"),
    route_mode: str = Form("THROUGH_HEAD"),
    head_user_id: str = Form(""),
    board_user_id: str = Form(""),
    his_batch_id: str = Form(""),
    cashbook_batch_id: str = Form(""),
    accounting_batch_id: str = Form(""),
):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    head_id = int(head_user_id) if str(head_user_id or "").strip() else None
    his_id = int(his_batch_id) if str(his_batch_id or "").strip() else None
    cashbook_id = int(cashbook_batch_id) if str(cashbook_batch_id or "").strip() else None
    accounting_id = int(accounting_batch_id) if str(accounting_batch_id or "").strip() else None

    voucher_id = save_cash_control_voucher_record(
        request,
        route_mode=route_mode,
        head_user_id=head_id,
        voucher_date=voucher_date,
        his_batch_id=his_id,
        cashbook_batch_id=cashbook_id,
        accounting_batch_id=accounting_id,
    )

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)


@router.post("/vouchers/{voucher_id}/basic/update")
async def cash_control_voucher_basic_update(
    request: Request,
    voucher_id: int,
    voucher_date: str = Form(""),
    sign_mode: str = Form("INTERNAL_DIGITAL"),
    route_mode: str = Form("THROUGH_HEAD"),
    head_user_id: str = Form(""),
    board_user_id: str = Form(""),
    his_batch_id: str = Form(""),
    cashbook_batch_id: str = Form(""),
    accounting_batch_id: str = Form(""),
):
    denied = require_login(request)
    if denied:
        return denied

    data = fetch_full_cash_control_voucher(voucher_id)
    voucher = data.get("voucher")
    if not voucher:
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    user = request.state.user
    if voucher["created_by"] != user["id"] and user.get("role_code") != "ADMIN":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    if voucher["status"] != "DRAFT":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    head_id = int(head_user_id) if str(head_user_id or "").strip() else None
    his_id = int(his_batch_id) if str(his_batch_id or "").strip() else None
    cashbook_id = int(cashbook_batch_id) if str(cashbook_batch_id or "").strip() else None
    accounting_id = int(accounting_batch_id) if str(accounting_batch_id or "").strip() else None

    update_cash_control_voucher_basic_record(
        request,
        voucher_id=voucher_id,
        route_mode=route_mode,
        head_user_id=head_id,
        voucher_date=voucher_date,
        his_batch_id=his_id,
        cashbook_batch_id=cashbook_id,
        accounting_batch_id=accounting_id,
    )

    return RedirectResponse(
        f"/cash-control?main_tab=voucher&editing_voucher_id={voucher_id}"
        f"&his_batch_id={his_id or ''}"
        f"&cashbook_batch_id={cashbook_id or ''}"
        f"&accounting_batch_id={accounting_id or ''}",
        status_code=303,
    )

@router.get("/vouchers/{voucher_id}")
def cash_control_voucher_detail(request: Request, voucher_id: int):
    denied = require_login(request)
    if denied:
        return denied

    data = fetch_full_cash_control_voucher(voucher_id)
    voucher = data.get("voucher")
    if not can_view_cash_control_voucher(request.state.user, voucher):
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    if request.state.user.get("role_code") in BOARD_ROLES:
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}/print", status_code=303)

    return templates.TemplateResponse(
        "cash_control_voucher_detail.html",
        {
            "request": request,
            "user": request.state.user,
            "cash_control_status_labels": CASH_CONTROL_STATUS_LABELS,
            **load_cash_control_signature_users(),
            **data,
        },
    )


@router.get("/vouchers/{voucher_id}/print")
def cash_control_voucher_print(request: Request, voucher_id: int):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    data = fetch_full_cash_control_voucher(voucher_id)
    voucher = data.get("voucher")
    if not can_view_cash_control_voucher(user, voucher):
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    if user.get("role_code") in BOARD_ROLES and voucher["status"] == "SUBMITTED_TO_BOARD":
        with get_conn() as conn:
            current_voucher = conn.execute(
                "SELECT status, current_handler, board_user_id FROM cash_control_vouchers WHERE id = ?",
                (voucher_id,),
            ).fetchone()

            if current_voucher and current_voucher["status"] == "SUBMITTED_TO_BOARD" and (
                current_voucher["current_handler"] == user["id"]
                or current_voucher["board_user_id"] == user["id"]
                or current_voucher["current_handler"] is None
            ):
                conn.execute(
                    "UPDATE cash_control_vouchers SET status = 'BOARD_VIEWED', current_handler = NULL WHERE id = ?",
                    (voucher_id,),
                )
                conn.execute(
                    """
                    INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, note)
                    VALUES (?, 'HDTV_XEM_PHIEU', ?, ?)
                    """,
                    (voucher_id, user["id"], "HĐTV đã mở xem Phiếu kiểm soát thu, chi tiền mặt."),
                )
                conn.commit()
                data = fetch_full_cash_control_voucher(voucher_id)
                voucher = data.get("voucher")

    # Chỉ phục vụ bản in: bảo đảm dòng "Thu tiền khám chữa bệnh"
    # tại cột "Ban KSNB kiểm tra" lấy đúng tổng Tiền mặt HIS của batch HIS gắn với phiếu.
    # Không sửa dữ liệu gốc đã lưu trong DB.
    his_batch_id = voucher["his_batch_id"] if voucher and "his_batch_id" in voucher.keys() else None
    if his_batch_id:
        with get_conn() as conn:
            his_cash_row = conn.execute(
                """
                SELECT COALESCE(SUM(tien_mat), 0) AS his_cash_amount
                FROM cash_control_employee_summaries
                WHERE batch_id = ?
                """,
                (his_batch_id,),
            ).fetchone()

        his_cash_amount = money_value(his_cash_row["his_cash_amount"] if his_cash_row else 0)

        for item in data.get("revenue_items", []):
            if item.get("item_name") == "Thu tiền khám chữa bệnh":
                old_checked_amount = money_value(item.get("ksnb_checked_amount"))
                item["ksnb_checked_amount"] = his_cash_amount
                item["ksnb_check_type"] = (
                    "DIFFERENT"
                    if abs(his_cash_amount - money_value(item.get("document_amount"))) >= 1
                    else "MATCH"
                )

                data["ksnb_total"] = money_value(data.get("ksnb_total")) - old_checked_amount + his_cash_amount
                data["diff_total"] = money_value(data.get("document_total")) - money_value(data.get("ksnb_total"))
                break

    return templates.TemplateResponse(
        "cash_control_voucher_print.html",
        {
            "request": request,
            "user": request.state.user,
            "format_vn_sqlite_datetime": format_vn_sqlite_datetime,
            "cash_control_status_labels": CASH_CONTROL_STATUS_LABELS,
            **data,
        },
    )


@router.post("/vouchers/{voucher_id}/delete")
def cash_control_voucher_delete(request: Request, voucher_id: int):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user

    with get_conn() as conn:
        voucher = conn.execute(
            "SELECT * FROM cash_control_vouchers WHERE id = ?",
            (voucher_id,),
        ).fetchone()

        if not voucher:
            return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

        if voucher["created_by"] != user["id"] and user.get("role_code") != "ADMIN":
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        if voucher["status"] != "DRAFT":
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        conn.execute(
            "DELETE FROM cash_control_voucher_routes WHERE voucher_id = ?",
            (voucher_id,),
        )
        conn.execute(
            "DELETE FROM cash_control_voucher_signatures WHERE voucher_id = ?",
            (voucher_id,),
        )
        conn.execute(
            "DELETE FROM cash_control_voucher_items WHERE voucher_id = ?",
            (voucher_id,),
        )
        conn.execute(
            "DELETE FROM cash_control_vouchers WHERE id = ?",
            (voucher_id,),
        )
        conn.commit()

    return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)


@router.post("/vouchers/{voucher_id}/items/update")
async def cash_control_voucher_items_update(request: Request, voucher_id: int):
    denied = require_login(request)
    if denied:
        return denied

    data = fetch_full_cash_control_voucher(voucher_id)
    voucher = data.get("voucher")
    if not voucher:
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    user = request.state.user
    if voucher["created_by"] != user["id"] and user.get("role_code") != "ADMIN":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    if voucher["status"] != "DRAFT":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    form = await request.form()

    expected_his_context = load_latest_his_context(voucher["his_batch_id"] if voucher["his_batch_id"] else None)
    expected_cashbook_context = load_cashbook_context(voucher["cashbook_batch_id"] if voucher["cashbook_batch_id"] else None)
    expected_accounting_context = load_accounting_context(voucher["accounting_batch_id"] if voucher["accounting_batch_id"] else None)
    expected_control_context = build_control_voucher_context(
        expected_his_context["totals"],
        expected_cashbook_context,
        expected_accounting_context,
    )
    expected_revenue_map = {
        row["item_name"]: row
        for row in expected_control_context.get("control_revenue_rows", [])
    }

    with get_conn() as conn:
        conn.execute(
            """
            DELETE FROM cash_control_voucher_items
            WHERE voucher_id = ?
              AND item_name = 'Thu tiền Quầy kính'
            """,
            (voucher_id,),
        )

        items = conn.execute(
            "SELECT * FROM cash_control_voucher_items WHERE voucher_id = ? ORDER BY group_type, item_order",
            (voucher_id,),
        ).fetchall()
        for item in items:
            item_id = item["id"]
            check_type = str(form.get(f"check_type_{item_id}") or "MATCH").strip()
            checked_amount_raw = str(form.get(f"checked_amount_{item_id}") or "").strip()
            note = str(form.get(f"note_{item_id}") or "").strip()

            if check_type not in {"MATCH", "DIFFERENT"}:
                check_type = "MATCH"

            group_type = str(item["group_type"] or "").strip().upper()

            if group_type == "REVENUE":
                expected_row = expected_revenue_map.get(str(item["item_name"] or "").strip())
                if expected_row:
                    checked_amount = money_value(expected_row.get("ksnb_amount"))
                    check_type = "DIFFERENT" if abs(money_value(expected_row.get("diff_amount"))) >= 1 else "MATCH"
                    if not note:
                        note = str(expected_row.get("note") or "").strip()
                else:
                    checked_amount = money_value(checked_amount_raw) if checked_amount_raw else money_value(item["ksnb_checked_amount"])
            else:
                if check_type == "DIFFERENT":
                    checked_amount = money_value(checked_amount_raw) if checked_amount_raw else 0.0
                else:
                    checked_amount = None

            conn.execute(
                """
                UPDATE cash_control_voucher_items
                SET ksnb_check_type = ?, ksnb_checked_amount = ?, ksnb_note = ?
                WHERE id = ? AND voucher_id = ?
                """,
                (check_type, checked_amount, note, item_id, voucher_id),
            )

        section_iv_result = str(form.get("section_iv_result") or "").strip()
        section_iv_note = str(form.get("section_iv_note") or "").strip()
        section_v_text = str(form.get("section_v_text") or "").strip()
        section_vi_text = str(form.get("section_vi_text") or "").strip()

        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET section_iv_result = ?, section_iv_note = ?, section_v_text = ?, section_vi_text = ?
            WHERE id = ?
            """,
            (section_iv_result, section_iv_note, section_v_text, section_vi_text, voucher_id),
        )

        rows = conn.execute(
            "SELECT * FROM cash_control_voucher_items WHERE voucher_id = ?",
            (voucher_id,),
        ).fetchall()
        item_dicts = [dict(row) for row in rows]
        revenue_items = [row for row in item_dicts if row["group_type"] == "REVENUE"]
        expense_items = [row for row in item_dicts if row["group_type"] == "EXPENSE"]

        revenue_document_total = sum(money_value(row["document_amount"]) for row in revenue_items)
        revenue_ksnb_total = sum(cash_control_item_check_amount(row) for row in revenue_items)
        expense_document_total = sum(money_value(row["document_amount"]) for row in expense_items)
        expense_ksnb_total = sum(cash_control_item_check_amount(row) for row in expense_items)

        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET revenue_document_total = ?,
                revenue_ksnb_total = ?,
                revenue_diff_total = ?,
                expense_document_total = ?,
                expense_ksnb_total = ?,
                expense_diff_total = ?
            WHERE id = ?
            """,
            (
                revenue_document_total,
                revenue_ksnb_total,
                revenue_document_total - revenue_ksnb_total,
                expense_document_total,
                expense_ksnb_total,
                expense_document_total - expense_ksnb_total,
                voucher_id,
            ),
        )
        conn.commit()

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)


@router.post("/vouchers/{voucher_id}/submit-no-signature")
def cash_control_voucher_submit_no_signature(
    request: Request,
    voucher_id: int,
    no_sign_head_user_id: str = Form(...),
):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    head_user_id_text = str(no_sign_head_user_id or "").strip()
    if not head_user_id_text:
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    try:
        head_user_id_value = int(head_user_id_text)
    except ValueError:
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    with get_conn() as conn:
        voucher = conn.execute(
            "SELECT * FROM cash_control_vouchers WHERE id = ?",
            (voucher_id,),
        ).fetchone()
        if not voucher or voucher["created_by"] != user["id"] or voucher["status"] != "DRAFT":
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        head_user = conn.execute(
            """
            SELECT *
            FROM users
            WHERE id = ?
              AND is_active = 1
              AND role_code IN ('TRUONG_BAN_KSNB','PHO_TRUONG_BAN_KSNB')
            """,
            (head_user_id_value,),
        ).fetchone()
        if not head_user:
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        conn.execute(
            "DELETE FROM cash_control_voucher_signatures WHERE voucher_id = ?",
            (voucher_id,),
        )
        conn.execute(
            """
            INSERT INTO cash_control_voucher_signatures(
                voucher_id, signer_id, signer_name, signer_role, signature_text
            )
            VALUES (?, ?, ?, 'NGUOI_KIEM_SOAT', ?)
            """,
            (voucher_id, user["id"], user["full_name"], cash_control_no_signature_text(user)),
        )
        conn.execute(
            """
            INSERT INTO cash_control_voucher_signatures(
                voucher_id, signer_id, signer_name, signer_role, signature_text
            )
            VALUES (?, ?, ?, 'TRUONG_BAN', ?)
            """,
            (
                voucher_id,
                head_user["id"],
                head_user["full_name"],
                cash_control_no_signature_text(head_user),
            ),
        )
        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET status = 'NO_SIGNATURE_INTERNAL',
                submitted_at = CURRENT_TIMESTAMP,
                submitted_to_board_at = NULL,
                board_user_id = NULL,
                current_handler = NULL
            WHERE id = ?
            """,
            (voucher_id,),
        )
        conn.execute(
            """
            INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
            VALUES (?, 'KHONG_KY_SO_LUU_NOI_BO', ?, ?, ?)
            """,
            (
                voucher_id,
                user["id"],
                head_user["id"],
                "Không ký số; in tên người kiểm soát và Trưởng/Phó Ban KSNB trên Phiếu. Phiếu lưu nội bộ, không trình HĐTV.",
            ),
        )
        conn.commit()

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)



@router.post("/vouchers/{voucher_id}/submit")
def cash_control_voucher_submit(
    request: Request,
    voucher_id: int,
    board_user_id: str = Form(""),
):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user

    with get_conn() as conn:
        voucher = conn.execute("SELECT * FROM cash_control_vouchers WHERE id = ?", (voucher_id,)).fetchone()
        if not voucher or voucher["created_by"] != user["id"] or voucher["status"] != "DRAFT":
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        board_user = None
        if voucher["route_mode"] == "DIRECT_BOARD":
            board_user_id_text = str(board_user_id or "").strip()
            if not board_user_id_text:
                return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)
            try:
                board_user_id_value = int(board_user_id_text)
            except ValueError:
                return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

            board_user = conn.execute(
                """
                SELECT *
                FROM users
                WHERE id = ?
                  AND is_active = 1
                  AND role_code IN ('TONG_GIAM_DOC','PHO_TGD_THUONG_TRUC','PHO_TONG_GIAM_DOC')
                """,
                (board_user_id_value,),
            ).fetchone()
            if not board_user:
                return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        conn.execute(
            """
            INSERT INTO cash_control_voucher_signatures(
                voucher_id, signer_id, signer_name, signer_role, signature_text
            )
            VALUES (?, ?, ?, 'NGUOI_KIEM_SOAT', ?)
            """,
            (voucher_id, user["id"], user["full_name"], cash_control_signature_text(user)),
        )

        if voucher["route_mode"] == "DIRECT_BOARD":
            if user["role_code"] in HEAD_ROLES:
                conn.execute(
                    """
                    INSERT INTO cash_control_voucher_signatures(
                        voucher_id, signer_id, signer_name, signer_role, signature_text
                    )
                    VALUES (?, ?, ?, 'TRUONG_BAN', ?)
                    """,
                    (
                        voucher_id,
                        user["id"],
                        user["full_name"],
                        cash_control_signature_text(user),
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO cash_control_voucher_signatures(
                        voucher_id, signer_id, signer_name, signer_role, signature_text
                    )
                    VALUES (?, ?, ?, 'TRUONG_BAN_UY_QUYEN', ?)
                    """,
                    (
                        voucher_id,
                        user["id"],
                        "Trưởng Ban Kiểm soát nội bộ",
                        "Đã ủy quyền. Đồng ý với nội dung kiểm tra",
                    ),
                )
            conn.execute(
                """
                UPDATE cash_control_vouchers
                SET status = 'SUBMITTED_TO_BOARD',
                    submitted_at = CURRENT_TIMESTAMP,
                    submitted_to_board_at = CURRENT_TIMESTAMP,
                    board_user_id = ?,
                    current_handler = ?
                WHERE id = ?
                """,
                (board_user["id"], board_user["id"], voucher_id),
            )
            conn.execute(
                """
                INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
                VALUES (?, 'TRINH_THANG_HDTV', ?, ?, ?)
                """,
                (voucher_id, user["id"], board_user["id"], "Người kiểm soát ký số nội bộ và trình thẳng HĐTV"),
            )
        else:
            conn.execute(
                """
                UPDATE cash_control_vouchers
                SET status = 'SUBMITTED_TO_HEAD',
                    submitted_at = CURRENT_TIMESTAMP,
                    current_handler = head_user_id
                WHERE id = ?
                """,
                (voucher_id,),
            )
            conn.execute(
                """
                INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
                VALUES (?, 'TRINH_TRUONG_BAN', ?, ?, ?)
                """,
                (voucher_id, user["id"], voucher["head_user_id"], "Người kiểm soát trình Trưởng/Phó Ban KSNB"),
            )

        conn.commit()

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)


@router.post("/vouchers/{voucher_id}/head-approve")
def cash_control_voucher_head_approve(
    request: Request,
    voucher_id: int,
    board_user_id: str = Form(...),
):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if user.get("role_code") not in HEAD_ROLES and user.get("role_code") != "ADMIN":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    board_user_id_text = str(board_user_id or "").strip()
    if not board_user_id_text:
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    try:
        board_user_id_value = int(board_user_id_text)
    except ValueError:
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    with get_conn() as conn:
        voucher = conn.execute("SELECT * FROM cash_control_vouchers WHERE id = ?", (voucher_id,)).fetchone()
        if not voucher or voucher["status"] != "SUBMITTED_TO_HEAD":
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        board_user = conn.execute(
            """
            SELECT *
            FROM users
            WHERE id = ?
              AND is_active = 1
              AND role_code IN ('TONG_GIAM_DOC','PHO_TGD_THUONG_TRUC','PHO_TONG_GIAM_DOC')
            """,
            (board_user_id_value,),
        ).fetchone()
        if not board_user:
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        conn.execute(
            """
            INSERT INTO cash_control_voucher_signatures(
                voucher_id, signer_id, signer_name, signer_role, signature_text
            )
            VALUES (?, ?, ?, 'TRUONG_BAN', ?)
            """,
            (voucher_id, user["id"], user["full_name"], cash_control_signature_text(user)),
        )
        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET status = 'SUBMITTED_TO_BOARD',
                submitted_to_board_at = CURRENT_TIMESTAMP,
                board_user_id = ?,
                current_handler = ?
            WHERE id = ?
            """,
            (board_user["id"], board_user["id"], voucher_id),
        )
        conn.execute(
            """
            INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, to_user_id, note)
            VALUES (?, 'TRUONG_BAN_TRINH_HDTV', ?, ?, ?)
            """,
            (
                voucher_id,
                user["id"],
                board_user["id"],
                "Trưởng/Phó Ban KSNB ký số nội bộ và trình Hội đồng thành viên.",
            ),
        )
        conn.commit()

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)


@router.post("/vouchers/{voucher_id}/board-save")
def cash_control_voucher_board_save(request: Request, voucher_id: int):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if user.get("role_code") not in BOARD_ROLES and user.get("role_code") != "ADMIN":
        return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

    with get_conn() as conn:
        voucher = conn.execute("SELECT * FROM cash_control_vouchers WHERE id = ?", (voucher_id,)).fetchone()
        if not voucher or voucher["status"] not in {"SUBMITTED_TO_BOARD", "BOARD_VIEWED"}:
            return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)

        conn.execute(
            """
            UPDATE cash_control_vouchers
            SET status = 'BOARD_SAVED',
                board_saved_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (voucher_id,),
        )
        conn.execute(
            """
            INSERT INTO cash_control_voucher_routes(voucher_id, action, from_user_id, note)
            VALUES (?, 'HDTV_LUU_PHIEU', ?, ?)
            """,
            (voucher_id, user["id"], "HĐTV đã xem và lưu Phiếu kiểm soát thu, chi tiền mặt"),
        )
        conn.commit()

    return RedirectResponse(f"/cash-control/vouchers/{voucher_id}", status_code=303)


@router.post("/accounting/import/{batch_id}/replace")
async def accounting_import_replace(request: Request, batch_id: int, accounting_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)

    if not accounting_file.filename:
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)

    ensure_cash_control_tables()

    original_name = safe_original_filename(accounting_file.filename)
    safe_name = f"accounting_replace_{batch_id}_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(accounting_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=accounting&error={str(ex)}", status_code=303)

    parsed = parse_accounting_excel(stored_path)

    with get_conn() as conn:
        batch = conn.execute("SELECT id FROM cash_control_accounting_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)

        conn.execute("DELETE FROM cash_control_accounting_summaries WHERE batch_id = ?", (batch_id,))
        conn.execute("DELETE FROM cash_control_accounting_employee_summaries WHERE batch_id = ?", (batch_id,))
        conn.execute(
            """
            UPDATE cash_control_accounting_batches
            SET original_filename = ?,
                created_by = ?,
                created_at = CURRENT_TIMESTAMP,
                accounting_non_cash_diff_total = ?
            WHERE id = ?
            """,
            (
                original_name,
                user["id"],
                money_value(parsed.get("accounting_non_cash_diff_total")),
                batch_id,
            ),
        )

        for item in parsed["summary_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_accounting_summaries(batch_id, item_name, amount, note)
                VALUES (?, ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["amount"], item.get("note", "")),
            )

        if parsed.get("refund_kcb_amount"):
            conn.execute(
                """
                INSERT INTO cash_control_accounting_summaries(batch_id, item_name, amount, note)
                VALUES (?, ?, ?, ?)
                """,
                (batch_id, "Chi hoàn tiền Khám chữa bệnh", parsed["refund_kcb_amount"], "Chi hoàn tiền KCB theo Bảng kê Kế toán"),
            )

        for item in parsed["employee_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_accounting_employee_summaries(
                    batch_id, employee_code, employee_name, accounting_cash_amount,
                    accounting_bank_ml_amount, accounting_qr_amount, accounting_pos_amount,
                    accounting_non_cash_amount, accounting_report_non_cash_amount,
                    accounting_non_cash_diff_amount, note
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    item["employee_code"],
                    item["employee_name"],
                    item["accounting_cash_amount"],
                    item.get("accounting_bank_ml_amount", 0),
                    item.get("accounting_qr_amount", 0),
                    item.get("accounting_pos_amount", 0),
                    item.get("accounting_non_cash_amount", 0),
                    item.get("accounting_report_non_cash_amount", 0),
                    item.get("accounting_non_cash_diff_amount", 0),
                    item.get("note", ""),
                ),
            )

        conn.commit()

    return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=accounting&accounting_batch_id={batch_id}", status_code=303)


@router.post("/accounting/import")
async def accounting_import(request: Request, accounting_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)

    if not accounting_file.filename:
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)

    ensure_cash_control_tables()

    original_name = safe_original_filename(accounting_file.filename)
    safe_name = f"accounting_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(accounting_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=accounting&error={str(ex)}", status_code=303)

    parsed = parse_accounting_excel(stored_path)

    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO cash_control_accounting_batches(
                title, original_filename, created_by, accounting_non_cash_diff_total
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                "Tổng hợp số báo cáo của Kế toán",
                original_name,
                user["id"],
                money_value(parsed.get("accounting_non_cash_diff_total")),
            ),
        )
        batch_id = cur.lastrowid

        for item in parsed["summary_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_accounting_summaries(batch_id, item_name, amount, note)
                VALUES (?, ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["amount"], item.get("note", "")),
            )

        if parsed.get("refund_kcb_amount"):
            conn.execute(
                """
                INSERT INTO cash_control_accounting_summaries(batch_id, item_name, amount, note)
                VALUES (?, ?, ?, ?)
                """,
                (batch_id, "Chi hoàn tiền Khám chữa bệnh", parsed["refund_kcb_amount"], "Chi hoàn tiền KCB theo Bảng kê Kế toán"),
            )

        for item in parsed["employee_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_accounting_employee_summaries(
                    batch_id, employee_code, employee_name, accounting_cash_amount,
                    accounting_bank_ml_amount, accounting_qr_amount, accounting_pos_amount,
                    accounting_non_cash_amount, accounting_report_non_cash_amount,
                    accounting_non_cash_diff_amount, note
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    item["employee_code"],
                    item["employee_name"],
                    item["accounting_cash_amount"],
                    item.get("accounting_bank_ml_amount", 0),
                    item.get("accounting_qr_amount", 0),
                    item.get("accounting_pos_amount", 0),
                    item.get("accounting_non_cash_amount", 0),
                    item.get("accounting_report_non_cash_amount", 0),
                    item.get("accounting_non_cash_diff_amount", 0),
                    item.get("note", ""),
                ),
            )

        conn.commit()

    return RedirectResponse("/cash-control?main_tab=summary&summary_tab=accounting", status_code=303)


@router.post("/cashbook/import/{batch_id}/replace")
async def cashbook_import_replace(request: Request, batch_id: int, cashbook_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)

    if not cashbook_file.filename:
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)

    ensure_cash_control_tables()

    original_name = safe_original_filename(cashbook_file.filename)
    safe_name = f"cashbook_replace_{batch_id}_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(cashbook_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=cashbook&error={str(ex)}", status_code=303)

    parsed = parse_cashbook_excel(stored_path)

    with get_conn() as conn:
        batch = conn.execute("SELECT id FROM cash_control_cashbook_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)

        conn.execute("DELETE FROM cash_control_cashbook_summaries WHERE batch_id = ?", (batch_id,))
        conn.execute("DELETE FROM cash_control_cashbook_entries WHERE batch_id = ?", (batch_id,))
        conn.execute(
            "UPDATE cash_control_cashbook_batches SET original_filename = ?, created_by = ?, created_at = CURRENT_TIMESTAMP WHERE id = ?",
            (original_name, user["id"], batch_id),
        )

        for item in parsed["receipt_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_summaries(
                    batch_id, summary_type, item_name, voucher_count, amount
                )
                VALUES (?, 'RECEIPT', ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["voucher_count"], item["amount"]),
            )

        for item in parsed["expense_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_summaries(
                    batch_id, summary_type, item_name, voucher_count, amount
                )
                VALUES (?, 'EXPENSE', ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["voucher_count"], item["amount"]),
            )

        for entry in parsed.get("detail_entries", []):
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_entries(
                    batch_id, doc_type, voucher_no, voucher_date, item_name,
                    debit_code, credit_code, description, amount
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    entry["doc_type"],
                    entry["voucher_no"],
                    entry["voucher_date"],
                    entry["item_name"],
                    entry["debit_code"],
                    entry["credit_code"],
                    entry["description"],
                    entry["amount"],
                ),
            )

        conn.commit()

    return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=cashbook&cashbook_batch_id={batch_id}", status_code=303)


@router.post("/cashbook/import")
async def cashbook_import(request: Request, cashbook_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)

    if not cashbook_file.filename:
        return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)

    ensure_cash_control_tables()

    original_name = safe_original_filename(cashbook_file.filename)
    safe_name = f"cashbook_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(cashbook_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=cashbook&error={str(ex)}", status_code=303)

    parsed = parse_cashbook_excel(stored_path)

    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO cash_control_cashbook_batches(title, original_filename, created_by)
            VALUES (?, ?, ?)
            """,
            ("Tổng hợp Sổ quỹ tiền mặt", original_name, user["id"]),
        )
        batch_id = cur.lastrowid

        for item in parsed["receipt_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_summaries(
                    batch_id, summary_type, item_name, voucher_count, amount
                )
                VALUES (?, 'RECEIPT', ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["voucher_count"], item["amount"]),
            )

        for item in parsed["expense_rows"]:
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_summaries(
                    batch_id, summary_type, item_name, voucher_count, amount
                )
                VALUES (?, 'EXPENSE', ?, ?, ?)
                """,
                (batch_id, item["item_name"], item["voucher_count"], item["amount"]),
            )

        for entry in parsed.get("detail_entries", []):
            conn.execute(
                """
                INSERT INTO cash_control_cashbook_entries(
                    batch_id, doc_type, voucher_no, voucher_date, item_name,
                    debit_code, credit_code, description, amount
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    entry["doc_type"],
                    entry["voucher_no"],
                    entry["voucher_date"],
                    entry["item_name"],
                    entry["debit_code"],
                    entry["credit_code"],
                    entry["description"],
                    entry["amount"],
                ),
            )

        conn.commit()

    return RedirectResponse("/cash-control?main_tab=summary&summary_tab=cashbook", status_code=303)


@router.post("/import/{batch_id}/replace")
async def cash_control_import_replace(request: Request, batch_id: int, revenue_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control", status_code=303)

    if not revenue_file.filename:
        return RedirectResponse("/cash-control", status_code=303)

    original_name = safe_original_filename(revenue_file.filename)
    safe_name = f"cash_control_replace_{batch_id}_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(revenue_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=his&error={str(ex)}", status_code=303)

    summaries = parse_revenue_excel(stored_path)

    with get_conn() as conn:
        batch = conn.execute("SELECT id FROM cash_control_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            return RedirectResponse("/cash-control", status_code=303)

        conn.execute("DELETE FROM cash_control_employee_summaries WHERE batch_id = ?", (batch_id,))
        conn.execute("DELETE FROM cash_control_his_entries WHERE batch_id = ?", (batch_id,))
        conn.execute(
            "UPDATE cash_control_batches SET original_filename = ?, created_by = ?, created_at = CURRENT_TIMESTAMP WHERE id = ?",
            (original_name, user["id"], batch_id),
        )

        for item in summaries.values():
            conn.execute(
                """
                INSERT INTO cash_control_employee_summaries(
                    batch_id, employee_code, employee_name, patient_count, txn_count,
                    tien_thu, tam_ung, tien_chi, mien_giam, huy_phieu,
                    thuc_thu, tien_mat, chuyen_khoan, qr, pos
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    item["employee_code"],
                    item["employee_name"],
                    item["patient_count"],
                    item["txn_count"],
                    item["tien_thu"],
                    item["tam_ung"],
                    item["tien_chi"],
                    item["mien_giam"],
                    item["huy_phieu"],
                    item["thuc_thu"],
                    item["tien_mat"],
                    item["chuyen_khoan"],
                    item["qr"],
                    item["pos"],
                ),
            )
            for detail in item.get("details", []):
                conn.execute(
                    """
                    INSERT INTO cash_control_his_entries(
                        batch_id, employee_code, employee_name, patient_code, patient_name,
                        tien_thu, tam_ung, tien_chi, huy_phieu, thuc_thu, tien_mat, chuyen_khoan, qr, pos
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        detail["employee_code"],
                        detail["employee_name"],
                        detail["patient_code"],
                        detail["patient_name"],
                        detail["tien_thu"],
                        detail["tam_ung"],
                        detail["tien_chi"],
                        detail["huy_phieu"],
                        detail["thuc_thu"],
                        detail["tien_mat"],
                        detail["chuyen_khoan"],
                        detail["qr"],
                        detail["pos"],
                    ),
                )
        conn.commit()

    return RedirectResponse(f"/cash-control/{batch_id}?main_tab=summary&summary_tab=his", status_code=303)


@router.post("/import")
async def cash_control_import(request: Request, revenue_file: UploadFile = File(...)):
    denied = require_login(request)
    if denied:
        return denied

    user = request.state.user
    if not can_create_voucher(user):
        return RedirectResponse("/cash-control", status_code=303)

    if not revenue_file.filename:
        return RedirectResponse("/cash-control", status_code=303)

    original_name = safe_original_filename(revenue_file.filename)
    safe_name = f"cash_control_{original_name}"
    stored_path = IMPORT_DIR / safe_name
    try:
        original_name = _save_cash_control_import_upload(revenue_file, stored_path)
    except ValueError as ex:
        return RedirectResponse(f"/cash-control?main_tab=summary&summary_tab=his&error={str(ex)}", status_code=303)

    summaries = parse_revenue_excel(stored_path)

    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO cash_control_batches(title, original_filename, report_date, created_by)
            VALUES (?, ?, '', ?)
            """,
            ("Kiểm soát thu chi tiền mặt", original_name, user["id"]),
        )
        batch_id = cur.lastrowid

        for item in summaries.values():
            conn.execute(
                """
                INSERT INTO cash_control_employee_summaries(
                    batch_id, employee_code, employee_name, patient_count, txn_count,
                    tien_thu, tam_ung, tien_chi, mien_giam, huy_phieu,
                    thuc_thu, tien_mat, chuyen_khoan, qr, pos
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    item["employee_code"],
                    item["employee_name"],
                    item["patient_count"],
                    item["txn_count"],
                    item["tien_thu"],
                    item["tam_ung"],
                    item["tien_chi"],
                    item["mien_giam"],
                    item["huy_phieu"],
                    item["thuc_thu"],
                    item["tien_mat"],
                    item["chuyen_khoan"],
                    item["qr"],
                    item["pos"],
                ),
            )
            for detail in item.get("details", []):
                conn.execute(
                    """
                    INSERT INTO cash_control_his_entries(
                        batch_id, employee_code, employee_name, patient_code, patient_name,
                        tien_thu, tam_ung, tien_chi, huy_phieu, thuc_thu, tien_mat, chuyen_khoan, qr, pos
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        detail["employee_code"],
                        detail["employee_name"],
                        detail["patient_code"],
                        detail["patient_name"],
                        detail["tien_thu"],
                        detail["tam_ung"],
                        detail["tien_chi"],
                        detail["huy_phieu"],
                        detail["thuc_thu"],
                        detail["tien_mat"],
                        detail["chuyen_khoan"],
                        detail["qr"],
                        detail["pos"],
                    ),
                )
        conn.commit()

    return RedirectResponse(f"/cash-control/{batch_id}?main_tab=summary&summary_tab=his", status_code=303)


@router.get("/{batch_id}")
def cash_control_detail(request: Request, batch_id: int):
    denied = require_login(request)
    if denied:
        return denied

    if request.state.user.get("role_code") in BOARD_ROLES:
        return RedirectResponse("/cash-control?main_tab=voucher", status_code=303)

    active_main_tab = request.query_params.get("main_tab") or "summary"
    active_summary_tab = request.query_params.get("summary_tab") or "his"

    with get_conn() as conn:
        selected_batch = conn.execute(
            """
            SELECT b.*, u.full_name AS creator_name
            FROM cash_control_batches b
            JOIN users u ON u.id = b.created_by
            WHERE b.id = ?
            """,
            (batch_id,),
        ).fetchone()

        if not selected_batch:
            return RedirectResponse("/cash-control", status_code=303)

        batches = conn.execute(
            """
            SELECT b.*, u.full_name AS creator_name
            FROM cash_control_batches b
            JOIN users u ON u.id = b.created_by
            ORDER BY b.id DESC
            """
        ).fetchall()

        summaries = conn.execute(
            """
            SELECT *,
                   COALESCE(chuyen_khoan, 0) + COALESCE(qr, 0) + COALESCE(pos, 0) AS khong_bang_tien_mat
            FROM cash_control_employee_summaries
            WHERE batch_id = ?
            ORDER BY employee_code, employee_name
            """,
            (batch_id,),
        ).fetchall()

    summary_dicts = [dict(row) for row in summaries]
    totals = compute_totals(summary_dicts)
    totals.setdefault("patient_count", 0)
    totals.setdefault("txn_count", 0)
    totals.setdefault("khong_bang_tien_mat", 0)

    selected_cashbook_batch_id = query_int(request, "cashbook_batch_id")
    selected_accounting_batch_id = query_int(request, "accounting_batch_id")

    cashbook_context = load_cashbook_context(selected_cashbook_batch_id)
    accounting_context = load_accounting_context(selected_accounting_batch_id)
    reconcile_context = build_reconcile_context(
        totals,
        summary_dicts,
        cashbook_context,
        accounting_context,
        batch_id,
    )
    control_voucher_context = build_control_voucher_context(
        totals,
        cashbook_context,
        accounting_context,
    )
    signature_user_context = load_cash_control_signature_users()

    return templates.TemplateResponse(
        "cash_control_index.html",
        {
            "request": request,
            "user": request.state.user,
            "batches": batches,
            "selected_batch": selected_batch,
            "summaries": summary_dicts,
            "totals": totals,
            "error": None,
            "active_main_tab": active_main_tab,
            "active_summary_tab": active_summary_tab,
            "cash_control_revenue_items": CASH_CONTROL_REVENUE_ITEMS,
            "cash_control_expense_items": CASH_CONTROL_EXPENSE_ITEMS,
            "format_vn_sqlite_datetime": format_vn_sqlite_datetime,
            "selected_his_batch_id": batch_id,
            "selected_cashbook_batch_id": selected_cashbook_batch_id or (cashbook_context["selected_cashbook_batch"]["id"] if cashbook_context["selected_cashbook_batch"] else None),
            **cashbook_context,
            **accounting_context,
            **reconcile_context,
            **control_voucher_context,
            **signature_user_context,
        },
    )